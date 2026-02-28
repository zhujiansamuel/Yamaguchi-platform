#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import re
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

# ご注文番号: W + 10 桁
ORDER_RE = re.compile(r"W\d{10}")  # 例: W1562921866


def extract_order_number_from_pdf(pdf_path: Path) -> str:
    """
    PDF から ご注文番号 (W**********) を抽出する
    1) 「ご注文番号」の行を優先
    2) 見つからない場合は全文から最初の W+10 桁を検索
    """
    texts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            texts.append(t)

    full_text = "\n".join(texts)

    # ご注文番号 の行を優先
    m = re.search(r"ご注.?文番号[:：]?\s*(W\d{10})", full_text)
    if m:
        return m.group(1)

    # 全文から W+10 桁を検索（フォールバック）
    m2 = ORDER_RE.search(full_text)
    if m2:
        return m2.group(0)

    raise ValueError("ご注文番号 (W+10 桁) を検出できませんでした。")


def sanitize_filename(name: str) -> str:
    """
    ファイル名に使えない記号を "_" に置き換える
    """
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def load_name_mapping_from_excel(xlsx_path: Path) -> dict:
    """
    Excel から変換マップを読み込む:
    キー: 「注文番号」
    値  : 「変更名前」
    期待ヘッダー: 1 行目に「注文番号」「変更名前」が存在する
    """
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise RuntimeError("Excel が空です。（ヘッダー行が見つかりません）")

    header = [str(c).strip() if c is not None else "" for c in header_row]

    def find_col(name: str):
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    idx_order = find_col("注文番号")
    idx_name = find_col("変更名前")

    if idx_order is None or idx_name is None:
        raise RuntimeError(
            f"Excel に「注文番号」または「変更名前」という列ヘッダーがありません。"
        )

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        order_raw = row[idx_order] if idx_order < len(row) else None
        name_raw = row[idx_name] if idx_name < len(row) else None

        order_no = str(order_raw).strip() if order_raw is not None else ""
        new_name = str(name_raw).strip() if name_raw is not None else ""
        if not order_no or not new_name:
            continue

        mapping[order_no] = new_name

    return mapping


def list_pdf_files(pdf_dir: Path) -> list[Path]:
    return sorted(pdf_dir.glob("*.pdf"))


def info(msg: str) -> None:
    """通常メッセージ → 標準出力（成功ログに表示）"""
    print(msg)


def warn(msg: str) -> None:
    """警告・スキップ情報 → 標準エラー（エラーログに表示）"""
    print(msg, file=sys.stderr)


def error(msg: str) -> None:
    """致命的エラー → 標準エラー"""
    print(msg, file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--folder", required=True, help="PDF の保存フォルダ")
    parser.add_argument("--excel", required=True, help="注文番号と変更名前の Excel ファイル")
    parser.add_argument("--filter", default="PDF", help="ログ表示用のフィルタ名（現在は PDF 固定）")
    parser.add_argument("--dry-run", action="store_true", help="確認のみ（ファイル名は実際には変更しない）")
    args = parser.parse_args()

    pdf_dir = Path(args.folder)
    excel_path = Path(args.excel)

    if not pdf_dir.is_dir():
        error(f"[エラー] フォルダが存在しません: {pdf_dir}")
        return 2
    if not excel_path.is_file():
        error(f"[エラー] Excel ファイルが存在しません: {excel_path}")
        return 3

    # 変換表を読み込み
    try:
        mapping = load_name_mapping_from_excel(excel_path)
    except Exception as e:
        error(f"[エラー] Excel の読み込みに失敗しました: {e}")
        return 4

    info(f"Excel から {len(mapping)} 件の変換ルールを読み込みました。")
    if not mapping:
        error("[エラー] 有効な変換ルールが 1 件もありません。")
        return 5

    pdf_files = list_pdf_files(pdf_dir)
    total = len(pdf_files)
    info(f"対象 PDF ファイル数: {total} 件（フィルタ: {args.filter}）")

    if total == 0:
        warn("[警告] 指定されたフォルダ内に PDF ファイルが見つかりませんでした。")
        # 進捗 100% を出しておく
        print("PROGRESS 100 処理中 0/0")
        sys.stdout.flush()
        return 0

    done = 0
    for pdf_path in pdf_files:
        try:
            order_no = extract_order_number_from_pdf(pdf_path)
        except Exception as e:
            warn(f"[スキップ] {pdf_path.name}: ご注文番号を取得できませんでした。理由: {e}")
            done += 1
            percent = int(done * 100 / total)
            print(f"PROGRESS {percent} 処理中 {done}/{total}")
            sys.stdout.flush()
            continue

        new_base = mapping.get(order_no)
        if not new_base:
            warn(
                f"[警告] 注文番号 {order_no} に対応する変更名前が Excel にありません。"
                f" ファイル: {pdf_path.name}"
            )
            done += 1
            percent = int(done * 100 / total)
            print(f"PROGRESS {percent} 処理中 {done}/{total}")
            sys.stdout.flush()
            continue

        safe_base = sanitize_filename(new_base)
        new_path = pdf_path.with_name(safe_base + pdf_path.suffix)

        if new_path.exists():
            warn(f"[警告] 変更後のファイル名が既に存在するためスキップしました: {new_path.name}")
            done += 1
            percent = int(done * 100 / total)
            print(f"PROGRESS {percent} 処理中 {done}/{total}")
            sys.stdout.flush()
            continue

        if args.dry_run:
            info(f"[確認のみ] {pdf_path.name} → {new_path.name}")
        else:
            try:
                pdf_path.rename(new_path)
                info(f"ファイル名を変更しました: {pdf_path.name} → {new_path.name}")
            except Exception as e:
                error(
                    f"[エラー] ファイル名の変更に失敗しました: {pdf_path.name} → {new_path.name}"
                    f" 理由: {e}"
                )

        done += 1
        percent = int(done * 100 / total)
        print(f"PROGRESS {percent} 処理中 {done}/{total}")
        sys.stdout.flush()

    info("すべての処理が完了しました。")
    return 0


if __name__ == "__main__":
    sys.exit(main())

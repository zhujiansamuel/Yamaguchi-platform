"""
Excel writer for generating Nextcloud Excel files.
Creates standardized Excel files with __id, __version, __op columns.
"""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    Writer for generating standardized Excel files with DATA sheet.

    Output format:
    Sheet "DATA":
    | __id | __version | __op | field1 | field2 | ... |
    """

    SHEET_NAME = 'DATA'
    REQUIRED_COLUMNS = ['__id', '__version', '__op']

    # Styling
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def __init__(self, field_names: List[str]):
        """
        Initialize writer with field names.

        Args:
            field_names: List of business field names (excluding __id, __version, __op)
        """
        self.field_names = field_names
        self.headers = self.REQUIRED_COLUMNS + field_names
        self.workbook = Workbook()

        # Remove default sheet and create DATA sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        self.sheet = self.workbook.create_sheet(self.SHEET_NAME, 0)
        self._write_headers()

    def _write_headers(self):
        """Write and style header row."""
        for col_idx, header in enumerate(self.headers, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # Auto-adjust column widths
        for col_idx, header in enumerate(self.headers, start=1):
            column_letter = get_column_letter(col_idx)
            # Set minimum width based on header length
            self.sheet.column_dimensions[column_letter].width = max(12, len(header) + 2)

    def add_row(self, row_data: Dict[str, Any]):
        """
        Add a data row to the Excel file.

        Args:
            row_data: Dictionary with field names as keys, must include __id, __version

        Example:
            {
                '__id': 123,
                '__version': datetime(2025, 1, 15, 10, 30, 0),
                '__op': None,
                'order_number': 'ORD-001',
                'confirmed_at': datetime(2025, 1, 15, 10, 0, 0),
                ...
            }
        """
        next_row = self.sheet.max_row + 1
        row_values = []

        for header in self.headers:
            value = row_data.get(header)

            # Format value for Excel
            if value is None:
                formatted_value = ''
            elif isinstance(value, datetime):
                formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, bool):
                formatted_value = 'TRUE' if value else 'FALSE'
            else:
                formatted_value = value

            row_values.append(formatted_value)

        # Write row
        for col_idx, value in enumerate(row_values, start=1):
            self.sheet.cell(row=next_row, column=col_idx, value=value)

    def add_rows(self, rows: List[Dict[str, Any]]):
        """
        Add multiple data rows.

        Args:
            rows: List of row dictionaries
        """
        for row_data in rows:
            self.add_row(row_data)

        logger.info(f"Added {len(rows)} rows to Excel file")

    def to_bytes(self) -> bytes:
        """
        Generate Excel file as bytes.

        Returns:
            Excel file content as bytes
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def close(self):
        """Close workbook."""
        if self.workbook:
            self.workbook.close()


class ExcelGenerator:
    """
    High-level Excel generator for model data.
    """

    @staticmethod
    def generate_from_queryset(queryset, field_mapping: Dict[str, str]) -> bytes:
        """
        Generate Excel file from Django queryset.

        Args:
            queryset: Django queryset
            field_mapping: Mapping of Django field names to Excel column names
                Example: {
                    'id': '__id',
                    'updated_at': '__version',
                    'order_number': 'order_number',
                    'confirmed_at': 'confirmed_at',
                }

        Returns:
            Excel file content as bytes
        """
        # Extract field names (excluding __id, __version, __op)
        field_names = [
            excel_name for django_name, excel_name in field_mapping.items()
            if excel_name not in ExcelWriter.REQUIRED_COLUMNS
        ]

        writer = ExcelWriter(field_names)

        try:
            for obj in queryset:
                row_data = {
                    '__id': obj.id,
                    '__version': obj.updated_at.isoformat() if hasattr(obj, 'updated_at') else datetime.now().isoformat(),
                    '__op': None,  # No operation for existing records
                }

                # Add business fields
                for django_field, excel_name in field_mapping.items():
                    if excel_name not in ExcelWriter.REQUIRED_COLUMNS:
                        value = getattr(obj, django_field, None)
                        row_data[excel_name] = value

                writer.add_row(row_data)

            return writer.to_bytes()

        finally:
            writer.close()

    @staticmethod
    def get_field_mapping(model_name: str) -> Dict[str, str]:
        """
        Get field mapping for a specific model.

        Args:
            model_name: Model name (e.g., 'Purchasing', 'OfficialAccount')

        Returns:
            Field mapping dictionary

        Raises:
            ValueError: If model not supported
        """
        mappings = {
            'Purchasing': {
                'id': '__id',
                'updated_at': '__version',
                'order_number': 'order_number',
                'confirmed_at': 'confirmed_at',
                'shipped_at': 'shipped_at',
                'estimated_website_arrival_date': 'estimated_website_arrival_date',
                'tracking_number': 'tracking_number',
                'estimated_delivery_date': 'estimated_delivery_date',
                'delivery_status': 'delivery_status',
                'account_used': 'account_used',
                'payment_method': 'payment_method',
            },
            'OfficialAccount': {
                'id': '__id',
                'updated_at': '__version',
                'account_id': 'account_id',
                'email': 'email',
                'name': 'name',
                'postal_code': 'postal_code',
                'address_line_1': 'address_line_1',
                'address_line_2': 'address_line_2',
                'address_line_3': 'address_line_3',
                'passkey': 'passkey',
            },
            'GiftCard': {
                'id': '__id',
                'updated_at': '__version',
                'card_number': 'card_number',
                'passkey1': 'passkey1',
                'passkey2': 'passkey2',
                'balance': 'balance',
            },
            'DebitCard': {
                'id': '__id',
                'updated_at': '__version',
                'card_number': 'card_number',
                'expiry_month': 'expiry_month',
                'expiry_year': 'expiry_year',
                'passkey': 'passkey',
                'balance': 'balance',
                'last_balance_update': 'last_balance_update',
            },
            'CreditCard': {
                'id': '__id',
                'updated_at': '__version',
                'card_number': 'card_number',
                'expiry_month': 'expiry_month',
                'expiry_year': 'expiry_year',
                'passkey': 'passkey',
                'credit_limit': 'credit_limit',
                'last_balance_update': 'last_balance_update',
            },
            'TemporaryChannel': {
                'id': '__id',
                'updated_at': '__version',
                'expected_time': 'expected_time',
                'record': 'record',
            },
        }

        if model_name not in mappings:
            raise ValueError(
                f"Unsupported model: {model_name}. "
                f"Supported models: {', '.join(mappings.keys())}"
            )

        return mappings[model_name]

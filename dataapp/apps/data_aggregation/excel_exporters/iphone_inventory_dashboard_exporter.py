"""
iPhone Inventory Dashboard Excel exporter.
iPhone库存Dashboard的Excel导出器。

This exporter creates a specialized Excel file for the Data_Dashboard folder
with the following features:
- Row 1: Variable names (hidden)
- Row 2: Japanese headers
- Row 3+: Data
- Cell protection: Only batch_level_1/2/3 columns are editable
- Password protection: Xdb73008762
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection, Font, Alignment, PatternFill, Border, Side
from django.apps import apps


class iPhoneInventoryDashboardExporter:
    """
    Excel exporter for iPhone Inventory Dashboard.
    Exports Inventory records where iphone field is not null,
    with related data from EcSite, Purchasing, OfficialAccount,
    LegalPersonOffline, and TemporaryChannel.
    """
    
    # Sheet protection password
    PROTECTION_PASSWORD = 'Xdb73008762'
    
    # Fields that are editable (not protected)
    EDITABLE_FIELDS = ['batch_level_1', 'batch_level_2', 'batch_level_3']
    
    # Field definitions: (field_name, japanese_header, source)
    # source: 'inventory', 'iphone', 'ecsite', 'purchasing', 'official_account', 
    #         'payment', 'legal_person', 'temp_channel'
    FIELD_DEFINITIONS = [
        # Basic Inventory fields
        ('iphone', 'iPhone機種', 'iphone'),
        ('imei', 'IMEI', 'inventory'),
        ('batch_level_1', 'バッチレベル1', 'inventory'),
        ('batch_level_2', 'バッチレベル2', 'inventory'),
        ('batch_level_3', 'バッチレベル3', 'inventory'),
        ('source1', 'ソース1', 'inventory'),
        ('source2', 'ソース2', 'inventory'),
        ('source3', 'ソース3', 'inventory'),
        ('source4', 'ソース4', 'inventory'),
        ('transaction_confirmed_at', '取引確認日時', 'inventory'),
        ('scheduled_arrival_at', '到着予定日時', 'inventory'),
        ('checked_arrival_at_1', '確認到着日時1', 'inventory'),
        ('checked_arrival_at_2', '確認到着日時2', 'inventory'),
        ('actual_arrival_at', '実際到着日時', 'inventory'),
        ('status', 'ステータス', 'inventory'),
        ('updated_at', '更新日時', 'inventory'),
        
        # EcSite (source1) fields
        ('ecsite_reservation_number', '予約番号', 'ecsite'),
        ('ecsite_username', 'ECサイトユーザー名', 'ecsite'),
        ('ecsite_method', '方法', 'ecsite'),
        ('ecsite_reservation_time', '予約日時', 'ecsite'),
        ('ecsite_visit_time', '訪問日時', 'ecsite'),
        ('ecsite_order_created_at', '注文作成日時', 'ecsite'),
        ('ecsite_info_updated_at', '情報更新日時', 'ecsite'),
        ('ecsite_order_detail_url', '注文詳細URL', 'ecsite'),
        
        # Purchasing (source2) fields
        ('purchasing_order_number', '注文番号', 'purchasing'),
        ('purchasing_official_account', '公式アカウント', 'purchasing'),
        ('purchasing_confirmed_at', '確認日時', 'purchasing'),
        ('purchasing_shipped_at', '発送日時', 'purchasing'),
        ('purchasing_estimated_website_arrival_date', '公式サイト到着予定日', 'purchasing'),
        ('purchasing_estimated_website_arrival_date_2', '公式サイト到着予定日2', 'purchasing'),
        ('purchasing_tracking_number', '追跡番号', 'purchasing'),
        ('purchasing_estimated_delivery_date', '配達予定日', 'purchasing'),
        ('purchasing_latest_delivery_status', '最新配達状況', 'purchasing'),
        ('purchasing_delivery_status_query_time', '配送状況照会日時', 'purchasing'),
        ('purchasing_delivery_status_query_source', '配送状況照会元', 'purchasing'),
        ('purchasing_official_query_url', '公式照会URL', 'purchasing'),
        ('purchasing_shipping_method', '配送方法', 'purchasing'),
        ('purchasing_last_info_updated_at', '最終情報更新日時', 'purchasing'),
        ('purchasing_payment_method', '支払方法', 'purchasing'),
        ('purchasing_account_used', '使用アカウント', 'purchasing'),
        ('purchasing_updated_at', '購入更新日時', 'purchasing'),
        ('purchasing_creation_source', '作成元', 'purchasing'),
        
        # OfficialAccount fields (via Purchasing)
        ('official_account_email', 'メールアドレス', 'official_account'),
        ('official_account_name', '名前', 'official_account'),
        ('official_account_postal_code', '郵便番号', 'official_account'),
        ('official_account_address_line_1', '住所1', 'official_account'),
        ('official_account_address_line_2', '住所2', 'official_account'),
        ('official_account_address_line_3', '住所3', 'official_account'),
        
        # Payment card numbers
        ('card_numbers', 'カード番号', 'payment'),
        
        # LegalPersonOffline (source3) fields
        ('legal_person_username', '法人ユーザー名', 'legal_person'),
        ('legal_person_appointment_time', '予約日時', 'legal_person'),
        ('legal_person_visit_time', '法人訪問日時', 'legal_person'),
        ('legal_person_updated_at', '法人更新日時', 'legal_person'),
        
        # TemporaryChannel (source4) fields
        ('temp_channel_created_time', '一時チャネル作成日時', 'temp_channel'),
        ('temp_channel_expected_time', '入庫予定日時', 'temp_channel'),
        ('temp_channel_record', '記録', 'temp_channel'),
        ('temp_channel_last_updated', '一時チャネル更新日時', 'temp_channel'),
    ]
    
    def __init__(self):
        self.Inventory = apps.get_model('data_aggregation', 'Inventory')
        self.GiftCardPayment = apps.get_model('data_aggregation', 'GiftCardPayment')
        self.DebitCardPayment = apps.get_model('data_aggregation', 'DebitCardPayment')
        self.CreditCardPayment = apps.get_model('data_aggregation', 'CreditCardPayment')
    
    def get_queryset(self):
        """
        Get the queryset for iPhone inventory records.
        Only returns records where iphone field is not null.
        """
        return self.Inventory.objects.filter(
            iphone__isnull=False,
            is_deleted=False
        ).select_related(
            'iphone',
            'source1',  # EcSite
            'source2',  # Purchasing
            'source2__official_account',  # OfficialAccount via Purchasing
            'source3',  # LegalPersonOffline
            'source4',  # TemporaryChannel
        ).prefetch_related(
            'source2__gift_card_payments__gift_card',
            'source2__debit_card_payments__debit_card',
            'source2__credit_card_payments__credit_card',
        )
    
    def format_iphone_info(self, iphone):
        """
        Format iPhone information as a combined string.
        Example: "iPhone 16 Pro 256GB ブラック"
        1024GB is displayed as 1T.
        """
        if not iphone:
            return ''
        
        model_name = iphone.model_name or ''
        capacity = iphone.capacity_gb
        color = iphone.color or ''
        
        # Format capacity: 1024GB -> 1T
        if capacity:
            if capacity >= 1024:
                capacity_str = f"{capacity // 1024}T"
            else:
                capacity_str = f"{capacity}GB"
        else:
            capacity_str = ''
        
        parts = [model_name, capacity_str, color]
        return ' '.join(p for p in parts if p)
    
    def format_datetime(self, dt):
        """
        Format datetime as string: 2025-01-11 10:30:00
        """
        if not dt:
            return ''
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    
    def format_date(self, d):
        """
        Format date as string: 2025-01-11
        """
        if not d:
            return ''
        return d.strftime('%Y-%m-%d')
    
    def format_source_indicator(self, source):
        """
        Format source field as indicator.
        Returns "✓" if source exists, empty string otherwise.
        """
        return '✓' if source else ''
    
    def get_card_numbers(self, purchasing):
        """
        Get all card numbers associated with a Purchasing instance.
        Returns card numbers joined by "｜".
        """
        if not purchasing:
            return ''
        
        card_numbers = []
        
        # Gift card payments
        for payment in purchasing.gift_card_payments.all():
            if payment.gift_card and payment.gift_card.card_number:
                card_numbers.append(payment.gift_card.card_number)
        
        # Debit card payments
        for payment in purchasing.debit_card_payments.all():
            if payment.debit_card and payment.debit_card.card_number:
                card_numbers.append(payment.debit_card.card_number)
        
        # Credit card payments
        for payment in purchasing.credit_card_payments.all():
            if payment.credit_card and payment.credit_card.card_number:
                card_numbers.append(payment.credit_card.card_number)
        
        return '｜'.join(card_numbers)
    
    def prepare_data(self):
        """
        Prepare and aggregate iPhone inventory data.
        This method is responsible for collecting and preparing all data.

        Returns:
            list: List of dictionaries, each representing one inventory record.
                  Each dictionary has field names as keys and formatted values as values.
        """
        queryset = self.get_queryset()
        data = []

        for inventory in queryset:
            record = {}
            for field_name, _, source in self.FIELD_DEFINITIONS:
                value = self._get_field_value(inventory, field_name, source)
                record[field_name] = value
            data.append(record)

        return data

    def get_row_data(self, inventory):
        """
        Get a row of data for an Inventory instance.
        Returns a list of values in the order of FIELD_DEFINITIONS.
        """
        row = []

        for field_name, _, source in self.FIELD_DEFINITIONS:
            value = self._get_field_value(inventory, field_name, source)
            row.append(value)

        return row
    
    def _get_field_value(self, inventory, field_name, source):
        """
        Get the value for a specific field.
        """
        if source == 'iphone':
            return self.format_iphone_info(inventory.iphone)
        
        elif source == 'inventory':
            if field_name in ['source1', 'source2', 'source3', 'source4']:
                return self.format_source_indicator(getattr(inventory, field_name))
            elif field_name in ['transaction_confirmed_at', 'scheduled_arrival_at', 
                               'checked_arrival_at_1', 'checked_arrival_at_2',
                               'actual_arrival_at', 'updated_at']:
                return self.format_datetime(getattr(inventory, field_name))
            else:
                return getattr(inventory, field_name) or ''
        
        elif source == 'ecsite':
            ecsite = inventory.source1
            if not ecsite:
                return ''
            attr_name = field_name.replace('ecsite_', '')
            value = getattr(ecsite, attr_name, '')
            if attr_name in ['reservation_time', 'visit_time', 'order_created_at', 'info_updated_at']:
                return self.format_datetime(value)
            return value or ''
        
        elif source == 'purchasing':
            purchasing = inventory.source2
            if not purchasing:
                return ''
            attr_name = field_name.replace('purchasing_', '')
            if attr_name == 'official_account':
                # Return indicator for official_account
                return self.format_source_indicator(purchasing.official_account)
            value = getattr(purchasing, attr_name, '')
            if attr_name in [
                'confirmed_at',
                'shipped_at',
                'last_info_updated_at',
                'updated_at',
                'delivery_status_query_time',
            ]:
                return self.format_datetime(value)
            elif attr_name in [
                'estimated_website_arrival_date',
                'estimated_website_arrival_date_2',
                'estimated_delivery_date',
            ]:
                return self.format_date(value)
            return value or ''
        
        elif source == 'official_account':
            purchasing = inventory.source2
            if not purchasing or not purchasing.official_account:
                return ''
            attr_name = field_name.replace('official_account_', '')
            return getattr(purchasing.official_account, attr_name, '') or ''
        
        elif source == 'payment':
            return self.get_card_numbers(inventory.source2)
        
        elif source == 'legal_person':
            legal_person = inventory.source3
            if not legal_person:
                return ''
            attr_name = field_name.replace('legal_person_', '')
            value = getattr(legal_person, attr_name, '')
            if attr_name in ['appointment_time', 'visit_time', 'updated_at']:
                return self.format_datetime(value)
            return value or ''
        
        elif source == 'temp_channel':
            temp_channel = inventory.source4
            if not temp_channel:
                return ''
            attr_name = field_name.replace('temp_channel_', '')
            value = getattr(temp_channel, attr_name, '')
            if attr_name in ['created_time', 'expected_time', 'last_updated']:
                return self.format_datetime(value)
            return value or ''
        
        return ''
    
    def export(self, existing_file_bytes=None):
        """
        Export iPhone inventory data to Excel format.
        
        Args:
            existing_file_bytes: If provided, write data to existing file.
                               If None, create a new file.
        
        Returns:
            io.BytesIO: Excel file as bytes stream
        """
        if existing_file_bytes:
            # Load existing workbook
            wb = load_workbook(filename=io.BytesIO(existing_file_bytes))
            ws = wb.active
            # Clear existing data (keep headers)
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'iPhone Inventory'
            
            # Write headers
            self._write_headers(ws)
        
        # Get data
        queryset = self.get_queryset()
        
        # Write data starting from row 3
        for row_idx, inventory in enumerate(queryset, start=3):
            row_data = self.get_row_data(inventory)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Apply protection to non-editable cells
                field_name = self.FIELD_DEFINITIONS[col_idx - 1][0]
                if field_name in self.EDITABLE_FIELDS:
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)
        
        # Apply sheet protection
        self._apply_protection(ws)
        
        # Hide row 1 (variable names)
        ws.row_dimensions[1].hidden = True
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _write_headers(self, ws):
        """
        Write header rows to worksheet.
        Row 1: Variable names (will be hidden)
        Row 2: Japanese headers
        """
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, (field_name, japanese_header, _) in enumerate(self.FIELD_DEFINITIONS, start=1):
            # Row 1: Variable name
            cell1 = ws.cell(row=1, column=col_idx, value=field_name)
            cell1.protection = Protection(locked=True)
            
            # Row 2: Japanese header
            cell2 = ws.cell(row=2, column=col_idx, value=japanese_header)
            cell2.font = header_font
            cell2.fill = header_fill
            cell2.alignment = header_alignment
            cell2.border = thin_border
            cell2.protection = Protection(locked=True)
    
    def _apply_protection(self, ws):
        """
        Apply sheet protection with password.
        Only batch_level_1/2/3 columns are editable.
        """
        ws.protection.sheet = True
        ws.protection.password = self.PROTECTION_PASSWORD
        ws.protection.enable()
    
    def _adjust_column_widths(self, ws):
        """
        Auto-adjust column widths based on content.
        """
        for col_idx, (field_name, japanese_header, _) in enumerate(self.FIELD_DEFINITIONS, start=1):
            column_letter = get_column_letter(col_idx)
            # Calculate width based on header length and field name
            max_length = max(len(field_name), len(japanese_header) * 2)  # Japanese chars are wider
            # Set minimum and maximum widths
            width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = width
    
    def get_field_names(self):
        """
        Get list of field names.
        """
        return [f[0] for f in self.FIELD_DEFINITIONS]
    
    def get_header_names(self):
        """
        Get mapping of field names to Japanese headers.
        """
        return {f[0]: f[1] for f in self.FIELD_DEFINITIONS}

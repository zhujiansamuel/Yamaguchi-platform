"""
Views for data_acquisition app.
"""
import logging
import os
import hmac
import hashlib
import json
import requests
from decimal import Decimal
from io import BytesIO
from datetime import datetime, date
from urllib.parse import unquote
from ipaddress import ip_address, ip_network

from openpyxl import load_workbook
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.apps import apps
from django.db import models

from rest_framework.decorators import api_view, authentication_classes, action
from rest_framework.response import Response
from rest_framework import status
from rest_framework.viewsets import ViewSet
from rest_framework.permissions import AllowAny
import uuid

from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiTypes

from apps.data_aggregation.authentication import SimpleTokenAuthentication
from apps.data_aggregation.excel_exporters import get_exporter
from apps.core.history import ChangeSourceContext, ChangeSource
from apps.data_aggregation.utils import _check_token, _resolve_source
from .tasks import (
    sync_nextcloud_excel,
    official_account_order_planning,
    process_webscraper_tracking,
    process_tracking_excel,      # 通用追踪任务处理函数
    process_japan_post_tracking_10_excel,  # Japan Post Tracking 10 合 1 特殊处理函数
    process_yamato_tracking_10_excel,  # Yamato Tracking 10 专用任务
    TRACKING_TASK_CONFIGS,       # 追踪任务配置字典
)
from .models import SyncLog
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
import httpx


logger = logging.getLogger(__name__)


@csrf_exempt
@require_http_methods(["POST"])
def nextcloud_webhook(request):
    """
    Webhook receiver for Nextcloud file change events.

    Validates webhook authenticity using token-based authentication,
    then dispatches Celery task for async processing.

    Request Headers:
        X-Nextcloud-Webhook-Token: Webhook authentication token

    Request Body (JSON):
        {
            "event": "file_changed",
            -----------------------------------------------------------------------------
            "path": "/data_platform/iPhone_test.xlsx",
            -----------------------------------------------------------------------------
            "user": "admin"
        }
    """
    # Get token from header
    token = request.headers.get('X-Nextcloud-Webhook-Token')
    
    if not token or token != settings.NEXTCLOUD_WEBHOOK_TOKEN:
        logger.warning(f"Unauthorized webhook attempt with token: {token}")
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=401)

    try:
        data = json.loads(request.body)
        event = data.get('event')
        file_path = data.get('path')
        user = data.get('user')

        if event == 'file_changed' and file_path:
            logger.info(f"File change detected via webhook: {file_path} by {user}")
            
            # Dispatch Celery task
            sync_nextcloud_excel.delay(file_path, user)
            
            return JsonResponse({'status': 'success', 'message': 'Task dispatched'})
        
        return JsonResponse({'status': 'ignored', 'message': 'Event not handled'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Webhook processing error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def onlyoffice_callback(request):
    """
    OnlyOffice document server callback receiver.

    This endpoint receives callbacks from OnlyOffice when documents are edited.
    It implements dual-callback architecture:
    1. Receives callback from OnlyOffice
    2. Processes data (parses Excel, syncs to database)
    3. Forwards callback to Nextcloud for file persistence

    Query Parameters:
        nextcloud_callback: Original Nextcloud callback URL (URL-encoded)
        file_path: Path of the file being edited
        user_id: Nextcloud user ID (optional)
        user_display_name: User's display name (optional)
        edit_start_time: ISO 8601 timestamp (optional)

    Request Body (JSON):
        OnlyOffice callback payload - see OnlyOffice API docs

    Returns:
        JSON response with error code
    """

    # Step 1: Validate IP address
    client_ip = get_client_ip(request)

    if not is_allowed_callback_ip(client_ip):
        logger.warning(
            f"OnlyOffice callback rejected: IP {client_ip} not in whitelist"
        )
        return JsonResponse({'error': 1}, status=403)

    # Step 2: Parse request
    try:
        # GET requests are health checks
        if request.method == 'GET':
            return JsonResponse({'status': 'healthy', 'error': 0})

        # Parse JSON body
        callback_data = json.loads(request.body.decode('utf-8'))

        # Extract query parameters
        nextcloud_callback = request.GET.get('nextcloud_callback', '')
        file_path = request.GET.get('file_path', '')
        user_id = request.GET.get('user_id', '')
        user_display_name = request.GET.get('user_display_name', '')
        edit_start_time = request.GET.get('edit_start_time', '')

        # Decode URL-encoded parameters
        if nextcloud_callback:
            nextcloud_callback = unquote(nextcloud_callback)

        if file_path:
            file_path = unquote(file_path)

    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        logger.error(f"OnlyOffice callback parse error: {e}")
        return JsonResponse({'error': 1}, status=400)

    # Step 3: Log callback receipt
    status = callback_data.get('status', -1)
    key = callback_data.get('key', '')
    url = callback_data.get('url', '')

    # Log full callback data for debugging
    logger.info(
        f"OnlyOffice callback received: status={status}, key={key}, "
        f"file={file_path}, user={user_id}, url_present={bool(url)}"
    )
    logger.debug(f"Full OnlyOffice callback data: {json.dumps(callback_data)}")

    SyncLog.objects.create(
        operation_type='onlyoffice_callback',
        file_path=file_path,
        message=f"OnlyOffice callback status={status}",
        details={
            'status': status,
            'key': key,
            'url': url,
            'user_id': user_id,
            'user_display_name': user_display_name,
            'edit_start_time': edit_start_time,
            'callback_data': callback_data,
        },
        success=True,
    )

    # Step 4: Process callback based on status
    # Status codes: https://api.onlyoffice.com/editors/callback
    # 2 = Document ready for saving
    if status == 2:
        logger.info(f"Processing OnlyOffice document save: {file_path}")

        try:
            # ============================================================================
            # Check if this file matches any tracking task configuration
            # 遍历所有追踪任务配置，检查文件路径和文件名是否匹配
            # 支持通过 Excel 文件触发的任务：
            # - official_website_redirect_to_yamato_tracking (旧名: process_official_website_redirect_to_yamato_tracking_excel)
            # - official_website_tracking
            # - yamato_tracking
            #
            # 注意：redirect_to_japan_post_tracking 不通过 Excel 触发，而是通过重定向逻辑动态创建
            # ============================================================================
            matched_task = None
            for task_name, config in TRACKING_TASK_CONFIGS.items():
                # 只匹配有 path_keyword 和 filename_prefix 的配置（即可通过 Excel 触发的任务）
                path_keyword = config.get('path_keyword')
                filename_prefix = config.get('filename_prefix')

                if (path_keyword and filename_prefix and
                    path_keyword in file_path and
                    os.path.basename(file_path).startswith(filename_prefix)):
                    matched_task = task_name
                    break

            if matched_task:
                config = TRACKING_TASK_CONFIGS[matched_task]

                # Always trigger the task - resume logic is handled in process_tracking_excel
                # The task will check existing TrackingJobs and skip already published URLs
                logger.info(f"Triggering {config['display_name']} task for: {file_path}")

                # Trigger the asynchronous task
                # Special handling for japan_post_tracking_10 (uses dedicated function)
                if matched_task == 'japan_post_tracking_10':
                    process_japan_post_tracking_10_excel.delay(
                        file_path=file_path,
                        document_url=url
                    )
                # yamato_tracking_10 使用专用的处理函数，其他任务使用通用函数
                elif matched_task == 'yamato_tracking_10':
                    process_yamato_tracking_10_excel.delay(
                        file_path=file_path,
                        document_url=url
                    )
                else:
                    process_tracking_excel.delay(
                        task_name=matched_task,
                        file_path=file_path,
                        document_url=url
                    )

                # Log the trigger event
                SyncLog.objects.create(
                    operation_type=config['sync_log_triggered'],
                    file_path=file_path,
                    message=f"{config['display_name']} task triggered",
                    success=True,
                    details={
                        'document_url': url,
                        'task_name': matched_task
                    }
                )
            # Download document
            elif url:
                # Option 1: Download from OnlyOffice provided URL
                process_onlyoffice_document(
                    document_url=url,
                    file_path=file_path,
                    callback_data=callback_data,
                    user_id=user_id
                )
            elif file_path:
                # Option 2: Fallback to WebDAV download if URL is missing but path is known
                logger.info(f"URL missing in callback, attempting fallback WebDAV download for: {file_path}")
                
                # Construct WebDAV URL from settings
                nc_config = settings.NEXTCLOUD_CONFIG
                base_url = nc_config['webdav_hostname'].rstrip('/')
                webdav_url = base_url + '/' + file_path.lstrip('/')
                auth = (nc_config['webdav_login'], nc_config['webdav_password'])
                
                logger.info(f"Fallback WebDAV URL: {webdav_url}")
                
                response = requests.get(webdav_url, auth=auth, timeout=30)
                if response.status_code == 200:
                    logger.info(f"Successfully downloaded file via WebDAV fallback: {file_path}")
                    sync_onlyoffice_excel_data(response.content, file_path, user_id)
                else:
                    logger.error(f"WebDAV fallback download failed: status={response.status_code}")
                    SyncLog.objects.create(
                        operation_type='onlyoffice_fallback_failed',
                        file_path=file_path,
                        message=f"WebDAV fallback failed with status {response.status_code}",
                        success=False,
                    )
            else:
                logger.warning(f"OnlyOffice callback missing both URL and file_path for status={status}")

        except Exception as e:
            logger.error(f"Failed to process OnlyOffice document: {e}", exc_info=True)
            SyncLog.objects.create(
                operation_type='onlyoffice_process_failed',
                file_path=file_path,
                message=f"Failed to process document: {e}",
                success=False,
                error_message=str(e),
            )

    # Step 5: Forward callback to Nextcloud
    if nextcloud_callback:
        try:
            forward_response = forward_callback_to_nextcloud(
                nextcloud_callback_url=nextcloud_callback,
                callback_data=callback_data
            )

            logger.info(
                f"Callback forwarded to Nextcloud: status={forward_response.status_code}"
            )

            # Return Nextcloud's response to OnlyOffice
            try:
                return JsonResponse(forward_response.json(), status=forward_response.status_code)
            except:
                return HttpResponse(forward_response.content, status=forward_response.status_code)

        except Exception as e:
            logger.error(f"Failed to forward callback to Nextcloud: {e}", exc_info=True)

            SyncLog.objects.create(
                operation_type='nextcloud_forward_failed',
                file_path=file_path,
                message=f"Failed to forward to Nextcloud: {e}",
                success=False,
                error_message=str(e),
            )

            # Still return success to OnlyOffice (we processed our part)
            return JsonResponse({'error': 0})

    # No Nextcloud callback URL provided
    return JsonResponse({'error': 0})


def get_client_ip(request):
    """Extract client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')

    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')

    return ip


def is_allowed_callback_ip(client_ip):
    """Check if client IP is in allowed callback IPs."""
    if not hasattr(settings, 'ALLOWED_CALLBACK_IPS'):
        return True  # No restriction if not configured

    allowed_ips = settings.ALLOWED_CALLBACK_IPS

    if not allowed_ips:
        return True  # Empty list means allow all

    from ipaddress import ip_address, ip_network
    try:
        client_ip_obj = ip_address(client_ip)

        for allowed in allowed_ips:
            allowed = allowed.strip()

            if not allowed:
                continue

            # Check if it's a network (CIDR notation)
            if '/' in allowed:
                if client_ip_obj in ip_network(allowed, strict=False):
                    return True
            # Check if it's a single IP
            elif client_ip_obj == ip_address(allowed):
                return True

        return False

    except ValueError as e:
        logger.error(f"Invalid IP address format: {client_ip} - {e}")
        return False


ONLYOFFICE_IMPORT_MODELS = {
    'iPhone',
    'iPad',
    'Inventory',
    'EcSite',
    'LegalPersonOffline',
    'TemporaryChannel',
    'OfficialAccount',
    'Purchasing',
    'GiftCard',
    'GiftCardPayment',
    'DebitCard',
    'DebitCardPayment',
    'CreditCard',
    'CreditCardPayment',
}


def get_onlyoffice_model_name(file_path):
    """
    Parse the model name from OnlyOffice file path.

    Expected naming: {ModelName}_test.xlsx or {ModelName}_test_{timestamp}.xlsx
    """
    import os
    import re

    filename = os.path.basename(file_path or '')
    name_part = filename.replace('.xlsx', '').replace('.xls', '')
    match = re.match(r'(?P<model>.+?)_test(?:_\d{8}_\d{6})?$', name_part)
    if match:
        return match.group('model')
    return None


def sync_onlyoffice_excel_data(excel_bytes, file_path, user_id):
    """
    Sync Excel data from OnlyOffice to database.
    """
    model_name = get_onlyoffice_model_name(file_path)
    if not model_name or model_name not in ONLYOFFICE_IMPORT_MODELS:
        logger.warning(f"Skipping sync for unknown model or file: {file_path}")
        return

    logger.info(f"Syncing {model_name} data from OnlyOffice Excel: {file_path}")

    # Use existing sync logic (similar to sync_nextcloud_excel task)
    from apps.data_aggregation.excel_exporters import get_exporter
    
    try:
        exporter = get_exporter(model_name)
        workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
        sheet = workbook.active

        # Get headers
        header_row = [cell.value for cell in sheet[1]]
        columns = build_excel_header_mapping(exporter, header_row)

        # Process rows
        rows_processed = 0
        model_class = exporter.get_model()
        
        # Get model field types to handle foreign keys
        model_fields = {f.name: f for f in model_class._meta.get_fields()}
        
        with ChangeSourceContext(ChangeSource.ONLYOFFICE_IMPORT):
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                data = {}
                for col_idx, field_name in columns:
                    raw_value = row[col_idx-1]
                    field = model_fields.get(field_name)
                    data[field_name] = convert_excel_value(raw_value, field)
                
                if not any(data.values()):
                    continue
                
                try:
                    # Update or create record
                    # We assume 'id' or 'card_number' or 'order_number' can be used as lookup
                    lookup_fields = ['id', 'card_number', 'order_number', 'account_id', 'uuid']
                    lookup_data = {}
                    update_data = {}
                    
                    for key, value in data.items():
                        # Skip read-only fields
                        if key in ['created_at', 'updated_at', 'last_info_updated_at']:
                            continue
                            
                        # Handle foreign keys: if field is 'official_account' but value is ID, 
                        # map it to 'official_account_id'
                        target_key = key
                        if key in model_fields:
                            field = model_fields[key]
                            if field.is_relation and not key.endswith('_id'):
                                # If it's a relation field and we have an ID-like value, use _id suffix
                                if isinstance(value, (int, str)) and str(value).isdigit():
                                    target_key = f"{key}_id"
                        
                        # Handle null values for non-nullable fields (like alternative_name)
                        if value is None:
                            if target_key in model_fields:
                                field = model_fields[target_key]
                                if not field.null:
                                    if isinstance(field, (models.CharField, models.TextField)):
                                        value = ""
                                    elif isinstance(field, (models.IntegerField, models.FloatField, models.DecimalField)):
                                        value = 0
                        
                        if target_key in lookup_fields and value is not None:
                            lookup_data[target_key] = value
                        else:
                            update_data[target_key] = value
                    
                    if lookup_data:
                        model_class.objects.update_or_create(
                            defaults=update_data,
                            **lookup_data
                        )
                        rows_processed += 1
                    else:
                        logger.warning(f"Row {row_idx}: No lookup fields found for {model_name}, skipping")
                except Exception as e:
                    logger.error(f"Error importing row {row_idx} for {model_name}: {e}")
                    SyncLog.objects.create(
                        operation_type='row_import_failed',
                        file_path=file_path,
                        message=f"Error importing row {row_idx}: {e}",
                        success=False,
                    )

        logger.info(f"Successfully synced {rows_processed} rows for {model_name}")

    except Exception as e:
        logger.error(f"Error syncing OnlyOffice data for {model_name}: {e}", exc_info=True)
        raise


def build_excel_header_mapping(exporter, header_row):
    header_names = exporter.get_header_names()
    reverse_mapping = {header: field for field, header in header_names.items()}
    columns = []

    for index, header in enumerate(header_row, start=1):
        if header is None:
            continue
        header_text = str(header).strip()
        if not header_text:
            continue
        field_name = reverse_mapping.get(header_text, header_text)
        columns.append((index, field_name))

    return columns


import pytz
from datetime import datetime

def convert_excel_value(value, field):
    if value is None:
        return None

    # Handle datetime values from Excel
    if isinstance(value, datetime):
        if value.tzinfo is None:
            # Assume naive datetime from Excel is Tokyo time
            tokyo_tz = pytz.timezone('Asia/Tokyo')
            value = tokyo_tz.localize(value)
        return value

    if isinstance(value, str):
        stripped = value.strip()
        if stripped == '':
            if isinstance(field, (models.CharField, models.TextField)):
                return ''
            return None
        value = stripped
    
    return value


def process_onlyoffice_document(document_url, file_path, callback_data, user_id):
    """
    Process OnlyOffice document after save.

    Downloads the document, parses Excel data, and syncs to database.
    """
    logger.info(f"Downloading document from OnlyOffice: {document_url}")

    try:
        # Download document
        response = requests.get(document_url, timeout=30)
        response.raise_for_status()

        sync_onlyoffice_excel_data(response.content, file_path, user_id)

        logger.info(f"Document processing complete: {file_path}")

        SyncLog.objects.create(
            operation_type='onlyoffice_document_processed',
            file_path=file_path,
            message=f"Document processed successfully",
            details={
                'document_url': document_url,
                'user_id': user_id,
                'file_size': len(response.content),
            },
            success=True,
        )

    except Exception as e:
        logger.error(f"Failed to process document: {e}", exc_info=True)
        raise


def forward_callback_to_nextcloud(nextcloud_callback_url, callback_data):
    """
    Forward OnlyOffice callback to Nextcloud.

    This ensures that Nextcloud receives the callback and saves the file properly.
    """
    logger.info(f"Forwarding callback to Nextcloud: {nextcloud_callback_url}")

    try:
        response = requests.post(
            nextcloud_callback_url,
            json=callback_data,
            headers={
                'Content-Type': 'application/json',
                'User-Agent': 'Django-OnlyOffice-Proxy/1.0'
            },
            timeout=30
        )

        logger.info(
            f"Nextcloud callback response: status={response.status_code}, "
            f"content={response.text[:200]}"
        )

        return response

    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to forward callback to Nextcloud: {e}", exc_info=True)
        raise


@csrf_exempt
@require_http_methods(["GET"])
def health_check(request):
    """
    Health check endpoint for Django backend.
    """
    return JsonResponse({'status': 'healthy', 'error': 0})



@extend_schema(
    tags=["Data Acquisition"],
    summary="Trigger Order Planning",
    description="Trigger order planning task for official accounts.",
    request=None,
    responses={
        200: {
            'type': 'object',
            'properties': {
                'status': {'type': 'string'},
                'message': {'type': 'string'}
            }
        },
        500: {
            'type': 'object',
            'properties': {
                'status': {'type': 'string'},
                'message': {'type': 'string'}
            }
        }
    }
)
@api_view(['POST'])
@authentication_classes([SimpleTokenAuthentication])
def order_planning(request):
    """
    Trigger order planning for official accounts.
    """
    try:
        official_account_order_planning.delay()
        return Response({'status': 'success', 'message': 'Order planning task dispatched'})
    except Exception as e:
        logger.error(f"Order planning dispatch error: {e}")
        return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _get_bool_param(request, name: str, default: bool = False) -> bool:
    """从请求中获取布尔参数"""
    val = request.query_params.get(name, None)
    if val is None and isinstance(getattr(request, "data", None), dict):
        val = request.data.get(name, None)
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in {"1", "true", "t", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "f", "no", "n", "off"}:
        return False
    return default




class WebScraperTrackingViewSet(ViewSet):
    """
    ViewSet for handling WebScraper webhook tracking requests.
    """
    
    @extend_schema(
        tags=["Resale / Price"],
        summary="WebScraper Webhook",
        parameters=[
            OpenApiParameter("dry_run", OpenApiTypes.BOOL, required=False),
            OpenApiParameter("t", OpenApiTypes.STR, required=False, description="短 token（可替代 token）"),
            OpenApiParameter("source", OpenApiTypes.STR, required=False),
        ],
        request=OpenApiTypes.BYTE,
        responses={202: OpenApiTypes.OBJECT, 200: OpenApiTypes.OBJECT}
    )
    @action(
        detail=False, 
        methods=["post"], 
        url_path="webscraper-tracking", 
        permission_classes=[AllowAny],
        parser_classes=[FormParser, MultiPartParser, JSONParser], 
    )
    def webscraper_tracking(self, request):
        """
        处理 WebScraper webhook 请求。
        接收来自 WebScraper 的 webhook 数据，验证 token然后异步处理 tracking 任务。
        """
        dry_run = str(request.query_params.get("dry_run") or "").lower() in {"1", "true", "t", "yes", "y"}
        dedupe = _get_bool_param(request, "dedupe", True)
        upsert = _get_bool_param(request, "upsert", False)
        bid = request.headers.get("X-Batch-Id") or request.query_params.get("batch_id") or request.data.get("batch_id")
        
        try:
            batch_uuid = uuid.UUID(str(bid)) if bid else uuid.uuid4()
        except Exception:
            batch_uuid = uuid.uuid4()

        # 验证 token
        if not _check_token(request, path_token=None):
            return Response({"detail": "Webhook token 不匹配"}, status=status.HTTP_403_FORBIDDEN)

        # 获取 job_id 和 source
        job_id = request.data.get("scrapingjob_id") or request.data.get("job_id") \
                 or request.query_params.get("scrapingjob_id") or request.query_params.get("job_id")
        source_name = _resolve_source(request)
        
        if not job_id or not source_name:
            return Response(
                {"detail": "Webhook 需要 job_id(scrapingjob_id) 与 source（或提供映射）"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 验证 tracker 是否存在
        try:
            from .trackers.registry_tracker import get_tracker
            tracker = get_tracker(source_name)
            logger.info(f"Found tracker for source: {source_name}")
        except KeyError as e:
            return Response(
                {"detail": f"未知任务: {source_name}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error getting tracker: {e}", exc_info=True)
            return Response(
                {"detail": f"获取 tracker 时出错: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



        # 调度 Celery 任务
        try:
            task = process_webscraper_tracking.delay(
                source_name=source_name,
                job_id=job_id,
                batch_uuid=str(batch_uuid),
                request_data=request.data if hasattr(request, 'data') else {},
                dry_run=dry_run,
                dedupe=dedupe,
                upsert=upsert
            )
            
            return Response({
                "mode": "webhook",
                "accepted": True,
                "task_id": task.id,
                "job_id": job_id,
                "source": source_name,
                "dry_run": dry_run,
                "dedupe": dedupe,
                "upsert": upsert,
                "batch_id": str(batch_uuid)
            }, status=202)
            
        except Exception as e:
            logger.error(f"Error dispatching tracking task: {e}", exc_info=True)
            return Response(
                {"detail": f"调度任务时出错: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        # return Response({"mode": "webhook", "accepted": True, "job_id": job_id,
        #                  "source": source_name, "dry_run": dry_run, "dedupe": dedupe, "upsert": upsert,
        #                  "batch_id": str(batch_uuid)}, status=202)

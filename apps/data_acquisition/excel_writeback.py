"""
Excel 回写工具模块

用于将 WebScraper 追踪数据回写到原始 Excel 文件中。
"""
import pandas as pd
import numpy as np
import logging
import requests
import io
import time
from openpyxl import load_workbook
from django.conf import settings

logger = logging.getLogger(__name__)


def safe_str(value):
    """将值安全转换为字符串，处理 NaN 和 None"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ''
    if pd.isna(value):
        return ''
    return str(value).strip() if value else ''


def extract_writeback_data(df: pd.DataFrame) -> str:
    """
    从 DataFrame 中提取需要回写的数据

    根据不同的列组合提取数据：
    1. 如果包含 "状態発生日"、"配送履歴"、"詳細1"、"取扱局"、"県名等"
       -> 提取"配送履歴"不为空的最后一行的这5个字段
    2. 如果包含 "data2"、"data3"、"name"
       -> 提取"data2"不为空的最后一行的这3个字段

    Args:
        df: WebScraper 返回的数据 DataFrame

    Returns:
        str: 用"｜｜｜"分隔的字符串，如果无法提取则返回空字符串
    """
    if df.empty:
        logger.warning("DataFrame is empty, cannot extract writeback data")
        return ''

    # 检查是否为 redirect_to_japan_post_tracking 数据
    japan_post_columns = ['状態発生日', '配送履歴', '詳細1', '取扱局', '県名等']
    if all(col in df.columns for col in japan_post_columns):
        # 找到"配送履歴"不为空的最后一行
        valid_rows = df[df['配送履歴'].notna() & (df['配送履歴'] != '')]

        if valid_rows.empty:
            logger.warning("No valid rows found with non-empty '配送履歴'")
            return ''

        last_row = valid_rows.iloc[-1]

        # 提取字段并用"｜｜｜"分隔
        values = [safe_str(last_row.get(col)) for col in japan_post_columns]
        result = '｜｜｜'.join(values)
        logger.info(f"Extracted japan_post data: {result}")
        return result

    # 检查是否为 official_website_redirect_to_yamato_tracking 数据
    yamato_columns = ['data2', 'data3', 'name']
    if all(col in df.columns for col in yamato_columns):
        # 找到"data2"不为空的最后一行
        valid_rows = df[df['data2'].notna() & (df['data2'] != '')]

        if valid_rows.empty:
            logger.warning("No valid rows found with non-empty 'data2'")
            return ''

        last_row = valid_rows.iloc[-1]

        # 提取字段并用"｜｜｜"分隔
        values = [safe_str(last_row.get(col)) for col in yamato_columns]
        result = '｜｜｜'.join(values)
        logger.info(f"Extracted yamato data: {result}")
        return result

    logger.warning(f"Unknown data format. Columns: {list(df.columns)}")
    return ''


def writeback_to_excel(
    file_path: str,
    row_index: int,
    writeback_data: str,
    task_id: str = '',
    max_retries: int = 3,
    retry_delay: float = 2.0
):
    """
    将数据回写到 Excel 文件的 C 列（带重试机制）

    Args:
        file_path: Nextcloud 文件路径
        row_index: Excel 中的行索引（从 0 开始，表头在第 0 行，数据从第 1 行开始）
        writeback_data: 要写入的数据（已用"｜｜｜"分隔）
        task_id: Celery 任务 ID（用于日志）
        max_retries: 最大重试次数（默认 3 次）
        retry_delay: 重试延迟（秒，使用指数退避）

    Returns:
        bool: 是否成功
    """
    if not writeback_data:
        logger.info(f"[Task {task_id}] No data to write back")
        return True  # 没有数据要写入，视为成功

    last_error = None

    for attempt in range(max_retries + 1):
        try:
            if attempt > 0:
                # 指数退避
                sleep_time = retry_delay * (2 ** (attempt - 1))
                logger.info(
                    f"[Task {task_id}] Retry attempt {attempt}/{max_retries}, "
                    f"waiting {sleep_time}s..."
                )
                time.sleep(sleep_time)

            # Step 1: 下载 Excel 文件
            logger.info(f"[Task {task_id}] Downloading Excel from: {file_path}")
            nc_config = settings.NEXTCLOUD_CONFIG
            base_url = nc_config['webdav_hostname'].rstrip('/')
            webdav_url = base_url + '/' + file_path.lstrip('/')
            auth = (nc_config['webdav_login'], nc_config['webdav_password'])

            response = requests.get(webdav_url, auth=auth, timeout=60)
            response.raise_for_status()
            content = response.content

            # Step 2: 加载 Excel 文件
            workbook = load_workbook(filename=io.BytesIO(content))
            sheet = workbook.active

            # Step 3: 计算 Excel 行号（row_index + 2，因为 Excel 从 1 开始，且第 1 行是表头）
            excel_row = row_index + 2

            # Step 4: 写入 C 列（列索引为 3）
            cell = sheet.cell(row=excel_row, column=3)
            cell.value = writeback_data

            logger.info(
                f"[Task {task_id}] Writing to cell C{excel_row}: {writeback_data[:50]}..."
            )

            # Step 5: 保存到内存
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            # Step 6: 上传回 Nextcloud
            logger.info(f"[Task {task_id}] Uploading modified Excel to: {file_path}")
            upload_response = requests.put(
                webdav_url,
                auth=auth,
                data=output.read(),
                timeout=60
            )
            upload_response.raise_for_status()

            logger.info(f"[Task {task_id}] Successfully wrote back to Excel at row {excel_row}")

            # 记录 SyncLog
            from .models import SyncLog
            SyncLog.objects.create(
                operation_type='excel_writeback',
                celery_task_id=task_id,
                file_path=file_path,
                message=f"Wrote data to cell C{excel_row} (attempt {attempt + 1})",
                success=True,
                details={
                    'row_index': row_index,
                    'excel_row': excel_row,
                    'writeback_data': writeback_data,
                    'attempt': attempt + 1
                }
            )

            return True

        except Exception as e:
            last_error = e
            error_msg = str(e)

            # 检查是否为文件锁定错误
            if 'is locked' in error_msg or 'lock' in error_msg.lower():
                logger.warning(
                    f"[Task {task_id}] File is locked (attempt {attempt + 1}/{max_retries + 1}): {error_msg}"
                )
                # 如果还有重试机会，继续循环
                if attempt < max_retries:
                    continue
            else:
                # 其他错误，直接记录并返回
                logger.error(
                    f"[Task {task_id}] Failed to write back to Excel: {e}",
                    exc_info=True
                )
                break

    # 所有重试都失败了
    logger.error(
        f"[Task {task_id}] Failed to write back to Excel after {max_retries + 1} attempts: {last_error}",
        exc_info=True
    )

    # 记录失败的 SyncLog
    from .models import SyncLog
    SyncLog.objects.create(
        operation_type='excel_writeback',
        celery_task_id=task_id,
        file_path=file_path,
        message=f"Failed to write back to Excel after {max_retries + 1} attempts",
        success=False,
        error_message=str(last_error),
        details={
            'row_index': row_index,
            'writeback_data': writeback_data,
            'max_retries': max_retries
        }
    )

    return False


def batch_writeback_to_excel(batch_uuid: str, task_id: str = ''):
    """
    批量回写：将已完成任务的数据写入 Excel

    此方法支持增量回写：
    - 每 10 个任务完成时触发一次
    - 批次完成时触发最后一次
    - 只写入有 writeback_data 的已完成任务

    Args:
        batch_uuid: TrackingBatch 的 UUID
        task_id: Celery 任务 ID（用于日志）

    Returns:
        dict: 回写结果统计
    """
    from .models import TrackingBatch, SyncLog

    try:
        # 查找 TrackingBatch
        batch = TrackingBatch.objects.get(batch_uuid=batch_uuid)

        file_path = batch.file_path
        logger.info(
            f"[Task {task_id}] Starting batch writeback for {batch_uuid}, "
            f"file: {file_path}"
        )

        # 收集所有已完成的 job 的回写数据
        writeback_data_map = {}  # {row_index: writeback_data}

        completed_jobs = batch.jobs.filter(status='completed').order_by('index')

        for job in completed_jobs:
            # 从 TrackingJob 读取已保存的回写数据
            if job.writeback_data:
                writeback_data_map[job.index] = job.writeback_data
                logger.info(
                    f"[Task {task_id}] Prepared writeback for job {job.custom_id}, "
                    f"row_index={job.index}, data={job.writeback_data[:30]}..."
                )
            else:
                logger.warning(
                    f"[Task {task_id}] Job {job.custom_id} has no writeback_data, skipping"
                )

        if not writeback_data_map:
            logger.info(f"[Task {task_id}] No data to write back for batch {batch_uuid}")
            return {
                'status': 'success',
                'total_jobs': 0,
                'written': 0
            }

        # 下载 Excel 文件
        logger.info(f"[Task {task_id}] Downloading Excel from: {file_path}")
        nc_config = settings.NEXTCLOUD_CONFIG
        base_url = nc_config['webdav_hostname'].rstrip('/')
        webdav_url = base_url + '/' + file_path.lstrip('/')
        auth = (nc_config['webdav_login'], nc_config['webdav_password'])

        response = requests.get(webdav_url, auth=auth, timeout=60)
        response.raise_for_status()
        content = response.content

        # 加载 Excel 文件
        workbook = load_workbook(filename=io.BytesIO(content))
        sheet = workbook.active

        # 批量写入所有数据
        written_count = 0
        for row_index, writeback_data in writeback_data_map.items():
            excel_row = row_index + 2  # 表头在第 1 行
            cell = sheet.cell(row=excel_row, column=3)
            cell.value = writeback_data
            written_count += 1
            logger.debug(
                f"[Task {task_id}] Wrote to cell C{excel_row}: {writeback_data[:30]}..."
            )

        logger.info(f"[Task {task_id}] Wrote {written_count} rows to Excel")

        # 保存到内存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # 上传回 Nextcloud
        logger.info(f"[Task {task_id}] Uploading modified Excel to: {file_path}")
        upload_response = requests.put(
            webdav_url,
            auth=auth,
            data=output.read(),
            timeout=60
        )
        upload_response.raise_for_status()

        logger.info(
            f"[Task {task_id}] Successfully completed batch writeback for {batch_uuid}, "
            f"wrote {written_count} rows"
        )

        # 更新 TrackingBatch 的回写完成时间
        from django.utils import timezone
        batch.writeback_completed_at = timezone.now()
        batch.save(update_fields=['writeback_completed_at'])

        # 记录 SyncLog
        SyncLog.objects.create(
            operation_type='excel_writeback',
            celery_task_id=task_id,
            file_path=file_path,
            message=f"Batch writeback completed: {written_count} rows written",
            success=True,
            details={
                'batch_uuid': str(batch_uuid),
                'total_jobs': len(writeback_data_map),
                'written': written_count,
                'row_indices': list(writeback_data_map.keys())
            }
        )

        return {
            'status': 'success',
            'total_jobs': len(writeback_data_map),
            'written': written_count
        }

    except TrackingBatch.DoesNotExist:
        logger.error(f"[Task {task_id}] TrackingBatch not found: {batch_uuid}")
        return {
            'status': 'error',
            'reason': 'batch_not_found'
        }

    except Exception as e:
        logger.error(
            f"[Task {task_id}] Batch writeback failed for {batch_uuid}: {e}",
            exc_info=True
        )

        # 记录失败的 SyncLog
        SyncLog.objects.create(
            operation_type='excel_writeback',
            celery_task_id=task_id,
            message=f"Batch writeback failed for {batch_uuid}",
            success=False,
            error_message=str(e),
            details={
                'batch_uuid': str(batch_uuid)
            }
        )

        return {
            'status': 'error',
            'reason': str(e)
        }

"""
Celery tasks for data_acquisition app.
All tasks here will be processed by the acquisition_queue worker.
"""
from .celery import app
import logging
import pandas as pd
import time
import requests
import io
from openpyxl import load_workbook
from django.conf import settings

logger = logging.getLogger(__name__)


# ============================================================================
# TRACKING TASK CONFIGURATIONS
# 配置说明：
# - path_keyword: Nextcloud 文件路径中必须包含的关键词
# - filename_prefix: 文件名必须以此前缀开头
# - api_token: WebScraper API Token（从 settings.WEBSCRAPER_API_TOKEN 读取）
# - sitemap_id: WebScraper Sitemap ID（从 settings.WEBSCRAPER_SITEMAP_IDS 读取）
# - custom_id_prefix: 创建爬虫任务时的 custom_id 前缀
# - sync_log_triggered: 触发时记录到 SyncLog 的 operation_type
# - sync_log_completed: 完成时记录到 SyncLog 的 operation_type
# ============================================================================

# 从 settings 获取配置，如果不存在则使用空值或默认值
WEBSCRAPER_API_TOKEN = getattr(settings, 'WEB_SCRAPER_API_TOKEN', '')
SITEMAP_IDS = getattr(settings, 'WEB_SCRAPER_SITEMAP_IDS', {
    'official_website_redirect_to_yamato_tracking': 1421177,
    'redirect_to_japan_post_tracking': 1422223,
    'official_website_tracking': 789,
    'yamato_tracking_only': 1423671,
    'japan_post_tracking_only': 1423655,
    'japan_post_tracking_10': 1424233,

})

TRACKING_TASK_CONFIGS = {
    'official_website_redirect_to_yamato_tracking': {
        'path_keyword': 'official_website_redirect_to_yamato_tracking',
        'filename_prefix': 'OWRYT-',
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('official_website_redirect_to_yamato_tracking', 1421177),
        'custom_id_prefix': 'owryt',
        'sync_log_triggered': 'official_website_redirect_to_yamato_tracking_triggered',
        'sync_log_completed': 'official_website_redirect_to_yamato_tracking_completed',
        'display_name': 'Official Website Redirect to Yamato Tracking',
    },
    'temporary_flexible_capture': {
        # NOTE: 此任务由 TemporaryFlexibleCaptureWorker 触发，复用 official_website_redirect_to_yamato_tracking 的 sitemap
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('official_website_redirect_to_yamato_tracking', 1421177),
        'custom_id_prefix': 'tfc',
        'sync_log_triggered': 'temporary_flexible_capture_triggered',
        'sync_log_completed': 'temporary_flexible_capture_completed',
        'display_name': 'Temporary Flexible Capture',
    },
    'redirect_to_japan_post_tracking': {
        # NOTE: 此任务不通过 Excel 文件触发，而是通过 official_website_redirect_to_yamato_tracking 的重定向逻辑动态创建
        # 因此不需要 path_keyword 和 filename_prefix
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('redirect_to_japan_post_tracking', 1422223),
        'custom_id_prefix': 'rtjpt',
        'sync_log_triggered': 'redirect_to_japan_post_tracking_triggered',
        'sync_log_completed': 'redirect_to_japan_post_tracking_completed',
        'display_name': 'Redirect to Japan Post Tracking',
    },
    'official_website_tracking': {
        'path_keyword': 'official_website_tracking',
        'filename_prefix': 'OWT-',
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('official_website_tracking', 789),
        'custom_id_prefix': 'owt',
        'sync_log_triggered': 'official_website_tracking_triggered',
        'sync_log_completed': 'official_website_tracking_completed',
        'display_name': 'Official Website Tracking',
    },
    'yamato_tracking_only': {
        'path_keyword': 'yamato_tracking_only',
        'filename_prefix': 'YTO-',
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('yamato_tracking_only', 1423671),
        'custom_id_prefix': 'yto',
        'sync_log_triggered': 'yamato_tracking_only_triggered',
        'sync_log_completed': 'yamato_tracking_only_completed',
        'display_name': 'Yamato Tracking Only',
        'url_template': 'http://jizen.kuronekoyamato.co.jp/jizen/servlet/crjz.b.NQ0010?id={tracking_number}',
    },
    'japan_post_tracking_only': {
        'path_keyword': 'japan_post_tracking_only',
        'filename_prefix': 'JPTO-',
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('japan_post_tracking_only', 1423655),
        'custom_id_prefix': 'jpto',
        'sync_log_triggered': 'japan_post_tracking_only_triggered',
        'sync_log_completed': 'japan_post_tracking_only_completed',
        'display_name': 'Japan Post Tracking Only',
        'url_template': 'https://trackings.post.japanpost.jp/services/srv/search?requestNo1={tracking_number}=&search.x=62&search.y=13&startingUrlPatten=&locale=ja',
    },
    'japan_post_tracking_10': {
        'path_keyword': 'japan_post_tracking_10',
        'filename_prefix': 'JPT10-',
        'api_token': WEBSCRAPER_API_TOKEN,
        'sitemap_id': SITEMAP_IDS.get('japan_post_tracking_10', 1424233),
        'custom_id_prefix': 'jpt10',
        'sync_log_triggered': 'japan_post_tracking_10_triggered',
        'sync_log_completed': 'japan_post_tracking_10_completed',
        'display_name': 'Japan Post Tracking 10',
    },
    'yamato_tracking_10': {
        'path_keyword': 'yamato_tracking_10',
        'filename_prefix': 'YT10-',
        'custom_id_prefix': 'yt10',
        'sync_log_triggered': 'yamato_tracking_10_triggered',
        'sync_log_completed': 'yamato_tracking_10_completed',
        'display_name': 'Yamato Tracking 10',
        # 注意：此任务不使用 WebScraper API，直接调用 query_yamato() 函数
    },

}


@app.task(name='apps.data_acquisition.tasks.sample_acquisition_task')
def sample_acquisition_task(source_url):
    """
    Sample task for data acquisition.
    This is a placeholder - implement your actual acquisition logic here.
    """
    logger.info(f"Processing acquisition task for URL: {source_url}")
    # Add your acquisition logic here
    return {"status": "completed", "url": source_url}


@app.task(name='apps.data_acquisition.tasks.fetch_data_from_source')
def fetch_data_from_source(source_config):
    """
    Fetch data from external source.
    This is a placeholder - implement your actual logic here.
    """
    logger.info(f"Fetching data with config: {source_config}")
    # Add your data fetching logic here
    return {"status": "completed", "config": source_config}


@app.task(
    name='apps.data_acquisition.tasks.sync_nextcloud_excel',
    bind=True,
    max_retries=3,
    default_retry_delay=60
)
def sync_nextcloud_excel(self, file_path, event_user='', event_time=None):
    """
    Sync Nextcloud Excel file to Django database.

    This task:
    1. Downloads Excel file via WebDAV
    2. Parses DATA sheet with __id, __version, __op columns
    3. Syncs data to database with version conflict detection
    4. Writes back __id for newly created records

    Args:
        file_path: Nextcloud file path (e.g., /Data/Purchasing_abc123.xlsx)
        event_user: Nextcloud user who triggered the event
        event_time: Event timestamp (ISO format string)

    Returns:
        Dictionary with sync results
    """
    from .sync_handler import SyncHandler
    from .models import SyncLog

    task_id = self.request.id
    logger.info(
        f"[Task {task_id}] Starting Nextcloud sync for {file_path} "
        f"(user: {event_user}, time: {event_time})"
    )

    try:
        # Create sync handler
        handler = SyncHandler(
            file_path=file_path,
            event_user=event_user,
            celery_task_id=task_id
        )

        # Execute sync
        result = handler.sync()

        logger.info(f"[Task {task_id}] Sync completed: {result}")
        return result

    except Exception as exc:
        logger.error(f"[Task {task_id}] Sync failed: {exc}", exc_info=True)

        # Log error
        SyncLog.objects.create(
            operation_type='sync_failed',
            celery_task_id=task_id,
            file_path=file_path,
            message=f"Sync task failed: {exc}",
            success=False,
            error_message=str(exc),
        )

        # Retry task if retries remaining
        if self.request.retries < self.max_retries:
            logger.info(f"[Task {task_id}] Retrying in {self.default_retry_delay}s...")
            raise self.retry(exc=exc)

        # Max retries reached
        raise


class OrderPlanningError(Exception):
    """Exception raised for order planning errors."""
    pass



@app.task(
    name='apps.data_acquisition.tasks.official_account_order_planning',
    bind=True,
)
def official_account_order_planning(
    self,
    batch_encoding,
    jan,
    inventory_count,
    cards_per_group,
    card_type
):
    """
    Official Account Order Planning task.
    官方账号订单规划任务。

    This task:
    1. Searches for OfficialAccounts with specific batch_encoding
    2. Groups cards (GiftCard/CreditCard/DebitCard) with specific batch_encoding
    3. Creates Purchasing orders with associated inventory items

    Args:
        batch_encoding (str): Batch encoding to filter accounts and cards
        jan (str): JAN code of the electronic product
        inventory_count (int): Number of inventory items per order
        cards_per_group (int): Number of cards per group
        card_type (str): Card type - 'GiftCard', 'CreditCard', or 'DebitCard'

    Returns:
        dict: Task result with status and details

    Raises:
        OrderPlanningError: If validation fails
    """
    from apps.data_aggregation.models import (
        OfficialAccount, GiftCard, CreditCard, DebitCard, Purchasing
    )

    task_id = self.request.id
    logger.info(
        f"[Task {task_id}] Starting Official Account Order Planning: "
        f"batch_encoding={batch_encoding}, jan={jan}, "
        f"inventory_count={inventory_count}, cards_per_group={cards_per_group}, "
        f"card_type={card_type}"
    )

    try:
        # Step 1: Validate cards_per_group
        # TODO: Support more than 3 cards per group in create_with_inventory
        if cards_per_group > 3:
            raise OrderPlanningError(
                f"cards_per_group ({cards_per_group}) exceeds maximum of 3. "
                "create_with_inventory currently supports up to 3 cards."
            )

        if cards_per_group < 1:
            raise OrderPlanningError(
                f"cards_per_group ({cards_per_group}) must be at least 1."
            )

        # Step 2: Search for OfficialAccounts with batch_encoding
        accounts = OfficialAccount.objects.filter(
            batch_encoding=batch_encoding
        ).order_by('email')  # Sort by email alphabetically

        account_count = accounts.count()
        logger.info(f"[Task {task_id}] Found {account_count} OfficialAccounts")

        if account_count == 0:
            raise OrderPlanningError(
                f"No OfficialAccount found with batch_encoding='{batch_encoding}'"
            )

        # Step 3: Search for cards with batch_encoding
        card_model_map = {
            'GiftCard': GiftCard,
            'CreditCard': CreditCard,
            'DebitCard': DebitCard,
        }

        if card_type not in card_model_map:
            raise OrderPlanningError(
                f"Invalid card_type: {card_type}. "
                f"Must be one of: {list(card_model_map.keys())}"
            )

        CardModel = card_model_map[card_type]
        cards = CardModel.objects.filter(batch_encoding=batch_encoding)

        card_count = cards.count()
        logger.info(f"[Task {task_id}] Found {card_count} {card_type} cards")

        if card_count == 0:
            raise OrderPlanningError(
                f"No {card_type} found with batch_encoding='{batch_encoding}'"
            )

        # Step 4: Validate card count is divisible by cards_per_group
        # TODO: Handle case where card count is not exactly divisible
        if card_count % cards_per_group != 0:
            raise OrderPlanningError(
                f"Card count ({card_count}) is not divisible by "
                f"cards_per_group ({cards_per_group}). "
                f"Remainder: {card_count % cards_per_group}"
            )

        # Step 5: Validate card groups count equals account count
        # TODO: Handle case where card groups count does not match account count
        card_groups_count = card_count // cards_per_group
        if card_groups_count != account_count:
            raise OrderPlanningError(
                f"Card groups count ({card_groups_count}) does not match "
                f"OfficialAccount count ({account_count}). "
                f"Each account needs exactly one card group."
            )

        # Step 6: Sort cards by alternative_name numeric suffix
        def extract_sort_key(card):
            """
            Extract numeric sort key from alternative_name.
            Format: XXXX-111 where 111 is the numeric part after '-'
            """
            alt_name = card.alternative_name or ''
            if '-' in alt_name:
                parts = alt_name.split('-')
                try:
                    return int(parts[-1])
                except (ValueError, IndexError):
                    return 0
            return 0

        sorted_cards = sorted(cards, key=extract_sort_key)
        logger.info(f"[Task {task_id}] Cards sorted by alternative_name")

        # Step 7: Group cards
        card_groups = []
        for i in range(0, len(sorted_cards), cards_per_group):
            group = sorted_cards[i:i + cards_per_group]
            card_groups.append(group)

        logger.info(f"[Task {task_id}] Created {len(card_groups)} card groups")

        # Step 8: Create Purchasing orders
        created_orders = []
        accounts_list = list(accounts)

        for idx, (account, card_group) in enumerate(zip(accounts_list, card_groups)):
            logger.info(
                f"[Task {task_id}] Creating order {idx + 1}/{account_count} "
                f"for account: {account.email}"
            )

            # Prepare create_with_inventory kwargs
            order_kwargs = {
                'email': account.email,
                'jan': jan,
                'inventory_count': inventory_count,
                'batch_encoding': batch_encoding,
            }

            # Add card numbers
            # TODO: Support passing payment_amount from request
            for card_idx, card in enumerate(card_group):
                card_num_key = f'card_number_{card_idx + 1}'
                payment_key = f'payment_amount_{card_idx + 1}'
                order_kwargs[card_num_key] = card.card_number
                order_kwargs[payment_key] = 0  # Default to 0

            # Create the order
            purchasing, inventories = Purchasing.create_with_inventory(**order_kwargs)

            created_orders.append({
                'purchasing_id': purchasing.id,
                'purchasing_uuid': purchasing.uuid,
                'order_number': purchasing.order_number,
                'account_email': account.email,
                'inventory_count': len(inventories),
                'card_numbers': [card.card_number for card in card_group],
            })

            logger.info(
                f"[Task {task_id}] Created Purchasing {purchasing.uuid} "
                f"with {len(inventories)} inventory items"
            )

        # Step 9: Return success result
        result = {
            'status': 'success',
            'task_id': task_id,
            'batch_encoding': batch_encoding,
            'jan': jan,
            'card_type': card_type,
            'accounts_processed': account_count,
            'orders_created': len(created_orders),
            'orders': created_orders,
        }

        logger.info(f"[Task {task_id}] Order planning completed successfully")
        return result

    except OrderPlanningError as e:
        logger.error(f"[Task {task_id}] Order planning validation error: {e}")
        return {
            'status': 'error',
            'task_id': task_id,
            'error_type': 'validation_error',
            'message': str(e),
        }
    except Exception as e:
        logger.error(f"[Task {task_id}] Order planning failed: {e}", exc_info=True)
        return {
            'status': 'error',
            'task_id': task_id,
            'error_type': 'unexpected_error',
            'message': str(e),
        }





@app.task(
    name='apps.data_acquisition.tasks.process_webscraper_tracking',
    bind=True,
    max_retries=3,
    default_retry_delay=60,
    queue='tracking_webhook_queue'
)
def process_webscraper_tracking(
    self,
    source_name: str,
    job_id: str,
    batch_uuid: str,
    request_data: dict,
    dry_run: bool = False,
    dedupe: bool = True,
    upsert: bool = False
):
    """
    处理 WebScraper webhook 的 tracking 任务。
    
    Args:
        source_name: Tracker 源名称
        job_id: WebScraper job ID
        batch_uuid: 批次 UUID
        request_data: 原始请求数据
        dry_run: 是否为试运行模式
        dedupe: 是否去重
        upsert: 是否使用 upsert 模式
    
    Returns:
        dict: 处理结果
    """
    from .trackers.registry_tracker import get_tracker
    from .models import TrackingJob

    task_id = self.request.id
    logger.info(
        f"[Task {task_id}] Processing webscraper tracking: "
        f"source={source_name}, job_id={job_id}, batch_uuid={batch_uuid}, "
        f"dry_run={dry_run}, dedupe={dedupe}, upsert={upsert}"
    )

    # ============================================================================
    # 提取 custom_id 并查找对应的 TrackingJob
    # custom_id 用于关联追踪任务，可以从多个地方获取
    # ============================================================================
    custom_id = request_data.get('custom_id', '')
    tracking_job = None

    if custom_id:
        try:
            tracking_job = TrackingJob.objects.get(custom_id=custom_id)
            logger.info(
                f"[Task {task_id}] Found TrackingJob for custom_id={custom_id}, "
                f"batch={tracking_job.batch.batch_uuid}"
            )
        except TrackingJob.DoesNotExist:
            logger.warning(
                f"[Task {task_id}] TrackingJob not found for custom_id={custom_id}. "
                f"This may be an old job created before tracking was implemented."
            )
    elif job_id:
        # 如果没有 custom_id，尝试通过 job_id 查找
        try:
            tracking_job = TrackingJob.objects.get(job_id=job_id)
            logger.info(
                f"[Task {task_id}] Found TrackingJob for job_id={job_id}"
            )
        except TrackingJob.DoesNotExist:
            logger.warning(
                f"[Task {task_id}] TrackingJob not found for job_id={job_id}"
            )

    try:
        # 获取 tracker
        tracker = get_tracker(source_name)
        
        # 如果是 dry_run，只记录不执行
        if dry_run:
            logger.info(f"[Task {task_id}] Dry run mode - skipping actual processing")
            return {
                'status': 'dry_run',
                'task_id': task_id,
                'source': source_name,
                'job_id': job_id,
                'batch_uuid': str(batch_uuid),
                'message': 'Dry run completed - no data processed'
            }
        

        from apps.data_aggregation.utils import fetch_webscraper_export_sync
        from .trackers.registry_tracker import run_tracker
        content = fetch_webscraper_export_sync(job_id, format="csv")
        df = pd.read_csv(io.BytesIO(content), encoding="utf-8-sig")

        # ============================================================================
        # 检查是否需要重定向到 redirect_to_japan_post_tracking
        # 如果 official_website_redirect_to_yamato_tracking 返回的数据中全部行的 Discriminant 列为空，
        # 则重定向到 redirect_to_japan_post_tracking
        # ============================================================================
        should_redirect = False
        if (source_name == 'official_website_redirect_to_yamato_tracking'
            and tracking_job
            and tracking_job.status != 'redirected'
            and len(df) > 0):
            # 检查全部行的 Discriminant 列是否都为空
            if 'Discriminant' in df.columns:
                # 检查所有行的 Discriminant 是否都为空（NaN、None、空字符串等）
                all_empty = True
                for idx, value in df['Discriminant'].items():
                    if not (pd.isna(value) or str(value).strip() == ''):
                        all_empty = False
                        break

                if all_empty:
                    should_redirect = True
                    logger.info(
                        f"[Task {task_id}] All rows have empty Discriminant column, "
                        f"will redirect to redirect_to_japan_post_tracking"
                    )
            else:
                logger.warning(
                    f"[Task {task_id}] Column 'Discriminant' not found in dataframe, "
                    f"columns: {list(df.columns)}"
                )

        if should_redirect:

            logger.info(
                f"[Task {task_id}] All rows have empty Discriminant for {source_name}, "
                f"redirecting to redirect_to_japan_post_tracking"
            )

            # 提取 URL
            target_url = None
            if 'web_start_url' in df.columns and len(df) > 0:
                target_url = df.iloc[0]['web_start_url']

            if not target_url:
                logger.error(f"[Task {task_id}] Cannot find web_start_url in CSV for redirect")
                # 继续执行原逻辑
            else:
                # 标记原任务为 redirected
                tracking_job.mark_redirected()
                logger.info(f"[Task {task_id}] Marked TrackingJob {tracking_job.custom_id} as redirected")

                # 记录 SyncLog
                from .models import SyncLog
                SyncLog.objects.create(
                    operation_type='official_website_redirect_to_yamato_tracking_completed',
                    celery_task_id=task_id,
                    message=f"Redirected to redirect_to_japan_post_tracking: {target_url}",
                    success=True,
                    details={
                        'original_custom_id': tracking_job.custom_id,
                        'redirected_url': target_url,
                        'reason': 'single_row_result'
                    }
                )

                # 获取 redirect_to_japan_post_tracking 配置
                jpt_config = TRACKING_TASK_CONFIGS['redirect_to_japan_post_tracking']
                API_TOKEN = jpt_config['api_token']
                SITEMAP_ID = jpt_config['sitemap_id']

                # 构建新的 custom_id
                new_custom_id = f"rtjpt-from-owryt-{tracking_job.custom_id}"

                try:
                    # 调用 WebScraper API 创建 redirect_to_japan_post_tracking 任务
                    payload = {
                        "sitemap_id": SITEMAP_ID,
                        "driver": "fulljs",
                        "page_load_delay": 2000,
                        "request_interval": 2000,
                        "start_urls": [target_url],
                        "custom_id": new_custom_id
                    }

                    api_url = f"https://api.webscraper.io/api/v1/scraping-job?api_token={API_TOKEN}"
                    api_response = requests.post(api_url, json=payload, timeout=30)

                    if api_response.status_code in [200, 201, 202]:
                        # 创建成功，创建新的 TrackingJob
                        response_data = api_response.json()
                        new_job_id = response_data.get('id') or response_data.get('job_id') or None

                        new_tracking_job = TrackingJob.objects.create(
                            batch=tracking_job.batch,
                            job_id=new_job_id,
                            custom_id=new_custom_id,
                            target_url=target_url,
                            index=tracking_job.index,
                            status='pending'
                        )

                        logger.info(
                            f"[Task {task_id}] Created redirect_to_japan_post_tracking job: "
                            f"job_id={new_job_id}, custom_id={new_custom_id}"
                        )

                        # 记录成功的 SyncLog
                        SyncLog.objects.create(
                            operation_type='redirect_to_japan_post_tracking_triggered',
                            celery_task_id=task_id,
                            message=f"Japan Post tracking triggered for redirected job: {new_custom_id}",
                            success=True,
                            details={
                                'original_custom_id': tracking_job.custom_id,
                                'new_custom_id': new_custom_id,
                                'new_job_id': new_job_id,
                                'target_url': target_url
                            }
                        )

                        # 返回成功结果（不执行原 tracker）
                        return {
                            'status': 'redirected',
                            'task_id': task_id,
                            'source': source_name,
                            'job_id': job_id,
                            'batch_uuid': str(batch_uuid),
                            'custom_id': custom_id,
                            'tracking_job_id': tracking_job.id,
                            'redirected_to': {
                                'custom_id': new_custom_id,
                                'job_id': new_job_id,
                                'tracking_job_id': new_tracking_job.id
                            },
                            'message': 'Redirected to redirect_to_japan_post_tracking'
                        }
                    else:
                        logger.error(
                            f"[Task {task_id}] Failed to create redirect_to_japan_post_tracking job: "
                            f"{api_response.status_code} - {api_response.text}"
                        )
                        # 创建失败的 SyncLog
                        SyncLog.objects.create(
                            operation_type='redirect_to_japan_post_tracking_triggered',
                            celery_task_id=task_id,
                            message=f"Failed to create redirect_to_japan_post_tracking job",
                            success=False,
                            error_message=f"API error: {api_response.status_code} - {api_response.text}",
                            details={
                                'original_custom_id': tracking_job.custom_id,
                                'target_url': target_url
                            }
                        )
                        # API 调用失败，返回 redirected 状态（原任务已标记为 redirected）
                        return {
                            'status': 'redirected_failed',
                            'task_id': task_id,
                            'source': source_name,
                            'job_id': job_id,
                            'batch_uuid': str(batch_uuid),
                            'custom_id': custom_id,
                            'tracking_job_id': tracking_job.id,
                            'message': 'Marked as redirected but failed to create redirect_to_japan_post_tracking job',
                            'error': f"API error: {api_response.status_code}"
                        }

                except Exception as redirect_exc:
                    logger.error(
                        f"[Task {task_id}] Exception while creating redirect_to_japan_post_tracking job: {redirect_exc}",
                        exc_info=True
                    )
                    # 创建失败的 SyncLog
                    SyncLog.objects.create(
                        operation_type='redirect_to_japan_post_tracking_triggered',
                        celery_task_id=task_id,
                        message=f"Exception while creating redirect_to_japan_post_tracking job",
                        success=False,
                        error_message=str(redirect_exc),
                        details={
                            'original_custom_id': tracking_job.custom_id,
                            'target_url': target_url
                        }
                    )
                    # 异常发生，返回 redirected 状态
                    return {
                        'status': 'redirected_failed',
                        'task_id': task_id,
                        'source': source_name,
                        'job_id': job_id,
                        'batch_uuid': str(batch_uuid),
                        'custom_id': custom_id,
                        'tracking_job_id': tracking_job.id,
                        'message': 'Marked as redirected but exception occurred',
                        'error': str(redirect_exc)
                    }

        # 执行原 tracker
        result = run_tracker(source_name, df)

        logger.info(f"[Task {task_id}] Tracking completed: {result}")

        # ============================================================================
        # 检查 tracker 执行结果
        # 如果 tracker 返回的消息包含错误关键词，认为执行失败
        # ============================================================================
        tracker_failed = False
        if isinstance(result, str):
            error_keywords = [
                'Failed to process',
                'Missing required columns',
                'Empty DataFrame',
                'No valid tracking data found',
                'Error updating record',
            ]
            for keyword in error_keywords:
                if keyword in result:
                    tracker_failed = True
                    logger.warning(f"[Task {task_id}] Tracker execution failed: {result}")
                    break

        # ============================================================================
        # 提取并保存回写数据到 TrackingJob
        # 无论是否启用即时回写，都需要提取并保存数据，供批量回写使用
        # ============================================================================
        if tracking_job:
            from .excel_writeback import extract_writeback_data

            # 提取需要回写的数据
            writeback_data = extract_writeback_data(df)

            if writeback_data:
                # 保存到 TrackingJob
                tracking_job.writeback_data = writeback_data
                tracking_job.save(update_fields=['writeback_data'])
                logger.info(
                    f"[Task {task_id}] Saved writeback data to TrackingJob: "
                    f"{writeback_data[:50]}..."
                )
            else:
                logger.info(f"[Task {task_id}] No data to write back")

        # ============================================================================
        # Excel 回写：将追踪数据回写到原始 Excel 文件
        #
        # 回写策略说明：
        # 1. 即时回写（IMMEDIATE_WRITEBACK=True）：每个任务完成后立即回写
        #    优点：实时性好，用户可以立即看到结果
        #    缺点：可能出现并发写入冲突（文件锁定错误）
        #
        # 2. 批量回写（IMMEDIATE_WRITEBACK=False，默认）：整个 batch 完成后一次性回写
        #    优点：避免并发冲突，性能更好
        #    缺点：需要等待整个 batch 完成才能看到结果
        #
        # 建议：使用批量回写（默认设置）
        # ============================================================================
        IMMEDIATE_WRITEBACK = getattr(settings, 'TRACKING_IMMEDIATE_WRITEBACK', False)

        if IMMEDIATE_WRITEBACK and tracking_job and writeback_data:
            from .excel_writeback import writeback_to_excel

            # 获取文件路径和行索引
            file_path = tracking_job.batch.file_path
            row_index = tracking_job.index

            logger.info(
                f"[Task {task_id}] Writing back to Excel (immediate mode): {file_path}, "
                f"row_index={row_index}, data={writeback_data[:50]}..."
            )

            # 执行回写（带重试）
            writeback_success = writeback_to_excel(
                file_path=file_path,
                row_index=row_index,
                writeback_data=writeback_data,
                task_id=task_id,
                max_retries=3,
                retry_delay=2.0
            )

            if writeback_success:
                logger.info(f"[Task {task_id}] Excel writeback completed successfully")
            else:
                logger.warning(
                    f"[Task {task_id}] Excel writeback failed after retries, "
                    f"will be handled by batch writeback"
                )
        elif tracking_job and writeback_data:
            logger.info(
                f"[Task {task_id}] Immediate writeback disabled, "
                f"data will be written back when batch completes"
            )

        # ============================================================================
        # 更新 TrackingJob 状态
        # 只有在 tracker 执行成功时才标记为已完成
        # 这会自动触发 TrackingBatch 的进度更新
        # ============================================================================
        if tracking_job:
            if tracker_failed:
                # Tracker 执行失败，标记为失败
                tracking_job.mark_failed(error_message=result)
                logger.warning(
                    f"[Task {task_id}] TrackingJob {tracking_job.custom_id} marked as failed due to tracker error"
                )
            else:
                # Tracker 执行成功，标记为完成
                tracking_job.mark_completed()
                logger.info(
                    f"[Task {task_id}] TrackingJob {tracking_job.custom_id} marked as completed. "
                    f"Batch progress: {tracking_job.batch.completed_jobs}/{tracking_job.batch.total_jobs} "
                    f"({tracking_job.batch.completion_percentage}%)"
                )

        return {
            'status': 'success',
            'task_id': task_id,
            'source': source_name,
            'job_id': job_id,
            'batch_uuid': str(batch_uuid),
            'custom_id': custom_id,
            'tracking_job_id': tracking_job.id if tracking_job else None,
            'result': result
        }

    except KeyError as e:
        error_msg = f"Unknown tracker: {e}"
        logger.error(f"[Task {task_id}] {error_msg}")

        # 更新 TrackingJob 状态为失败
        if tracking_job:
            tracking_job.mark_failed(error_message=error_msg)
            logger.info(f"[Task {task_id}] TrackingJob {tracking_job.custom_id} marked as failed")

        return {
            'status': 'error',
            'task_id': task_id,
            'error_type': 'unknown_tracker',
            'message': str(e),
            'custom_id': custom_id,
            'tracking_job_id': tracking_job.id if tracking_job else None,
        }
    except Exception as exc:
        error_msg = str(exc)
        logger.error(f"[Task {task_id}] Tracking failed: {exc}", exc_info=True)

        # Retry task if retries remaining
        if self.request.retries < self.max_retries:
            logger.info(f"[Task {task_id}] Retrying in {self.default_retry_delay}s...")
            raise self.retry(exc=exc)

        # Max retries reached - mark as failed
        if tracking_job:
            tracking_job.mark_failed(error_message=error_msg)
            logger.info(f"[Task {task_id}] TrackingJob {tracking_job.custom_id} marked as failed after max retries")

        return {
            'status': 'error',
            'task_id': task_id,
            'error_type': 'processing_error',
            'message': error_msg,
            'custom_id': custom_id,
            'tracking_job_id': tracking_job.id if tracking_job else None,
        }


@app.task(
    name='apps.data_acquisition.tasks.batch_writeback_tracking_data',
    bind=True,
    max_retries=2,
    default_retry_delay=60,
    queue='tracking_webhook_queue'
)
def batch_writeback_tracking_data(self, batch_uuid: str):
    """
    批量回写追踪数据到 Excel 文件

    当整个 TrackingBatch 完成时调用此任务，一次性将所有数据写回 Excel，
    避免并发写入冲突。

    Args:
        batch_uuid: TrackingBatch 的 UUID（字符串格式）

    Returns:
        dict: 回写结果
    """
    from .excel_writeback import batch_writeback_to_excel

    task_id = self.request.id
    logger.info(f"[Task {task_id}] Starting batch writeback for batch {batch_uuid}")

    try:
        result = batch_writeback_to_excel(batch_uuid=batch_uuid, task_id=task_id)

        logger.info(f"[Task {task_id}] Batch writeback completed: {result}")
        return result

    except Exception as exc:
        logger.error(
            f"[Task {task_id}] Batch writeback failed for {batch_uuid}: {exc}",
            exc_info=True
        )

        # Retry if retries remaining
        if self.request.retries < self.max_retries:
            logger.info(f"[Task {task_id}] Retrying in {self.default_retry_delay}s...")
            raise self.retry(exc=exc)

        # Max retries reached
        return {
            'status': 'error',
            'reason': str(exc),
            'max_retries_reached': True
        }


@app.task(
    name='apps.data_acquisition.tasks.publish_tracking_batch',
    bind=True,
    max_retries=0,  # 不重试，超时即抛弃
    time_limit=60,  # 1 分钟超时
    soft_time_limit=55,
    queue='publish_tracking_queue'
)
def publish_tracking_batch(self, task_name, url, batch_uuid_str, custom_id, index):
    """
    发布单个追踪任务到 WebScraper API（优化版）

    此函数只处理单个 URL 的发布，由 Excel 处理函数批量调用。
    执行完成后强制睡眠 6 秒，确保 API 频率限制。

    频率限制策略：
    - 每个任务完成后（无论成功、失败、跳过）都睡眠 6 秒
    - 这确保每 15 分钟最多发布约 150 个任务（900秒/6秒=150），远低于 200 个限制

    Args:
        task_name: 任务名称，必须是 TRACKING_TASK_CONFIGS 中的 key
        url: 单个 URL 字符串
        batch_uuid_str: TrackingBatch 的 UUID 字符串
        custom_id: 自定义任务 ID
        index: URL 在批次中的索引

    Returns:
        dict: {
            'status': 'success' | 'failed' | 'skipped',
            'custom_id': str,
            'job_id': str (如果成功),
            'url': str,
            'index': int
        }
    """
    from .models import SyncLog, TrackingBatch, TrackingJob

    # 获取任务配置
    if task_name not in TRACKING_TASK_CONFIGS:
        raise ValueError(f"Unknown task name: {task_name}")

    config = TRACKING_TASK_CONFIGS[task_name]
    task_id = self.request.id

    logger.info(
        f"[Task {task_id}] Publishing single URL: "
        f"task={task_name}, custom_id={custom_id}, url={url}"
    )

    result = None  # 用于存储返回结果

    try:
        # 查找 TrackingBatch
        tracking_batch = TrackingBatch.objects.filter(
            batch_uuid=batch_uuid_str
        ).first()

        if not tracking_batch:
            logger.error(f"[Task {task_id}] TrackingBatch not found: {batch_uuid_str}")
            result = {
                'status': 'failed',
                'custom_id': custom_id,
                'url': url,
                'index': index,
                'reason': 'batch_not_found'
            }
            return result

        # 检查是否已发布（断点续传）
        if TrackingJob.objects.filter(batch=tracking_batch, custom_id=custom_id).exists():
            logger.info(f"[Task {task_id}] URL already published, skipping: {custom_id}")
            result = {
                'status': 'skipped',
                'custom_id': custom_id,
                'url': url,
                'index': index
            }
            return result

        # 调用 WebScraper API
        API_TOKEN = config['api_token']
        SITEMAP_ID = config['sitemap_id']

        payload = {
            "sitemap_id": SITEMAP_ID,
            "driver": "fulljs",
            "page_load_delay": 2000,
            "request_interval": 2000,
            "start_urls": [url],
            "custom_id": custom_id
        }

        api_url = f"https://api.webscraper.io/api/v1/scraping-job?api_token={API_TOKEN}"
        api_response = requests.post(api_url, json=payload, timeout=30)

        if api_response.status_code in [200, 201, 202]:
            response_data = api_response.json()
            job_id = response_data.get('id') or response_data.get('job_id') or None

            # 创建 TrackingJob
            TrackingJob.objects.create(
                batch=tracking_batch,
                job_id=job_id,
                custom_id=custom_id,
                target_url=url,
                index=index,
                status='pending'
            )

            logger.info(
                f"[Task {task_id}] Successfully published: {custom_id} (job_id={job_id})"
            )

            result = {
                'status': 'success',
                'custom_id': custom_id,
                'job_id': job_id,
                'url': url,
                'index': index
            }
            return result
        else:
            logger.error(
                f"[Task {task_id}] WebScraper API error: "
                f"{api_response.status_code} - {api_response.text}"
            )
            result = {
                'status': 'failed',
                'custom_id': custom_id,
                'url': url,
                'index': index,
                'reason': f"API error: {api_response.status_code}"
            }
            return result

    except Exception as exc:
        logger.error(
            f"[Task {task_id}] Failed to publish {custom_id}: {exc}",
            exc_info=True
        )
        result = {
            'status': 'failed',
            'custom_id': custom_id,
            'url': url,
            'index': index,
            'reason': str(exc)
        }
        return result

    finally:
        # ============================================================================
        # 强制睡眠 6 秒（API 频率限制）
        # 无论成功、失败还是跳过，都睡眠 6 秒，确保不超过 API 频率限制
        # 计算：15 分钟 = 900 秒，900 / 6 = 150 个任务，远低于 200 个限制
        # ============================================================================
        time.sleep(6)
        logger.debug(f"[Task {task_id}] Rate limit sleep (6s) completed")


@app.task(
    name='apps.data_acquisition.tasks.process_tracking_excel',
    bind=True,
    max_retries=2,
    default_retry_delay=300,
    queue='tracking_excel_queue'
)
def process_tracking_excel(self, task_name, file_path, document_url=None):
    """
    处理 Excel 文件并投递追踪任务（优化版）

    新架构：
    1. 快速下载并解析 Excel 文件，提取所有 URLs
    2. 创建 TrackingBatch
    3. 批量投递任务到 publish_tracking_queue，每个任务间隔 2 秒
    4. Phase 1 任务快速完成，实际发布由 Phase 1.5 worker 处理

    Args:
        task_name: 任务名称
        file_path: Nextcloud 文件路径
        document_url: OnlyOffice 提供的下载 URL（可选）

    Returns:
        dict: {
            'status': 'success',
            'batch_uuid': str,
            'total_urls': int,
            'dispatched': int  # 投递的任务数
        }
    """
    from .models import SyncLog, TrackingBatch, TrackingJob
    import uuid

    # 获取任务配置
    if task_name not in TRACKING_TASK_CONFIGS:
        raise ValueError(f"Unknown task name: {task_name}")

    config = TRACKING_TASK_CONFIGS[task_name]
    task_id = self.request.id

    logger.info(
        f"[Task {task_id}] Starting {config['display_name']} Excel processing for {file_path}"
    )

    try:
        # Step 1: Download document
        content = None
        if document_url:
            logger.info(f"[Task {task_id}] Downloading from URL: {document_url}")
            response = requests.get(document_url, timeout=60)
            response.raise_for_status()
            content = response.content
        else:
            # Fallback to WebDAV
            logger.info(f"[Task {task_id}] Downloading via WebDAV: {file_path}")
            nc_config = settings.NEXTCLOUD_CONFIG
            base_url = nc_config['webdav_hostname'].rstrip('/')
            webdav_url = base_url + '/' + file_path.lstrip('/')
            auth = (nc_config['webdav_login'], nc_config['webdav_password'])

            response = requests.get(webdav_url, auth=auth, timeout=60)
            response.raise_for_status()
            content = response.content

        # Step 2: Parse Excel and extract URLs
        workbook = load_workbook(filename=io.BytesIO(content), data_only=False)
        sheet = workbook.active

        urls = []
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
            cell_a = row[0]
            cell_b = row[1] if len(row) > 1 else None
            url = None

            # Priority 1: Extract hyperlink from cell A
            if cell_a.hyperlink:
                url = cell_a.hyperlink.target
            # Priority 2: Check if cell A text is a URL
            elif isinstance(cell_a.value, str) and (cell_a.value.startswith('http://') or cell_a.value.startswith('https://')):
                url = cell_a.value
            # Priority 3: Construct Apple Store URL
            elif cell_a.value and cell_b and cell_b.value:
                cell_b_value = str(cell_b.value).strip()
                if '@' in cell_b_value:
                    cell_a_value = str(cell_a.value).strip()
                    url = f"https://store.apple.com/go/jp/vieworder/{cell_a_value}/{cell_b_value}"
            # Priority 4: Construct URL from url_template
            elif cell_a.value and config.get('url_template'):
                tracking_number = str(cell_a.value).strip()
                url = config['url_template'].format(tracking_number=tracking_number)

            if url:
                urls.append(url)

        logger.info(f"[Task {task_id}] Extracted {len(urls)} URLs from {file_path}")

        # Step 3: Create or find TrackingBatch
        batch, created = TrackingBatch.objects.get_or_create(
            file_path=file_path,
            task_name=task_name,
            defaults={
                'batch_uuid': uuid.uuid4(),
                'celery_task_id': task_id,
                'total_jobs': len(urls),
                'status': 'pending'
            }
        )

        batch_uuid_str = str(batch.batch_uuid)
        batch_short = batch_uuid_str[:8]

        if created:
            logger.info(f"[Task {task_id}] Created TrackingBatch {batch_short}")
        else:
            logger.info(f"[Task {task_id}] Found existing TrackingBatch {batch_short}")

        # Step 4: 批量投递任务到 publish_tracking_queue
        # 检查已投递的任务，只投递未投递的
        dispatched_count = 0
        skipped_count = 0

        for idx, url in enumerate(urls):
            custom_id = f"{config['custom_id_prefix']}-{batch_short}-{idx:04d}"

            # 检查是否已投递（断点续传）
            if TrackingJob.objects.filter(batch=batch, custom_id=custom_id).exists():
                skipped_count += 1
                logger.debug(f"[Task {task_id}] Skipping already dispatched: {custom_id}")
                continue

            # 投递任务到 publish_tracking_queue
            # 使用 countdown 参数实现 2 秒间隔
            publish_tracking_batch.apply_async(
                args=[task_name, url, batch_uuid_str, custom_id, idx],
                countdown=dispatched_count * 2  # 每个任务延迟 2 秒
            )

            dispatched_count += 1

        logger.info(
            f"[Task {task_id}] Dispatched {dispatched_count} tasks, skipped {skipped_count}"
        )

        # 记录 SyncLog
        SyncLog.objects.create(
            operation_type=config['sync_log_triggered'],
            celery_task_id=task_id,
            message=f"Dispatched {dispatched_count} publish tasks",
            success=True,
            details={
                'file_path': file_path,
                'batch_uuid': batch_uuid_str,
                'total_urls': len(urls),
                'dispatched': dispatched_count,
                'skipped': skipped_count
            }
        )

        return {
            'status': 'success',
            'batch_uuid': batch_uuid_str,
            'total_urls': len(urls),
            'dispatched': dispatched_count,
            'skipped': skipped_count
        }

    except Exception as exc:
        logger.error(
            f"[Task {task_id}] {config['display_name']} task failed: {exc}",
            exc_info=True
        )
        # Retry if retries remaining
        if self.request.retries < self.max_retries:
            raise self.retry(exc=exc)
        raise


@app.task(
    name='apps.data_acquisition.tasks.process_japan_post_tracking_10_excel',
    bind=True,
    max_retries=3,
    default_retry_delay=300,
    queue='tracking_excel_queue'
)
def process_japan_post_tracking_10_excel(self, file_path, document_url=None):
    """
    处理 Japan Post Tracking 10 合 1 追踪任务的 Excel 文件（优化版）

    新架构：
    1. 快速解析 Excel，提取追踪号并构造所有 URLs
    2. 创建 TrackingBatch
    3. 批量投递任务到 publish_tracking_queue，每个任务间隔 2 秒

    Args:
        file_path: Nextcloud 文件路径
        document_url: OnlyOffice 提供的下载 URL（可选）

    Returns:
        dict: {
            'status': 'success',
            'batch_uuid': str,
            'total_urls': int,
            'dispatched': int,
            'warnings': list
        }
    """
    from .models import SyncLog, TrackingBatch, TrackingJob
    import re
    import random
    import uuid
    from urllib.parse import urlencode

    task_name = 'japan_post_tracking_10'
    config = TRACKING_TASK_CONFIGS[task_name]
    task_id = self.request.id

    logger.info(
        f"[Task {task_id}] Starting {config['display_name']} Excel processing for {file_path}"
    )

    warnings = []

    try:
        # Step 1: Download document
        content = None
        if document_url:
            logger.info(f"[Task {task_id}] Downloading from URL: {document_url}")
            response = requests.get(document_url, timeout=60)
            response.raise_for_status()
            content = response.content
        else:
            # Fallback to WebDAV
            logger.info(f"[Task {task_id}] Downloading via WebDAV: {file_path}")
            nc_config = settings.NEXTCLOUD_CONFIG
            base_url = nc_config['webdav_hostname'].rstrip('/')
            webdav_url = base_url + '/' + file_path.lstrip('/')
            auth = (nc_config['webdav_login'], nc_config['webdav_password'])

            response = requests.get(webdav_url, auth=auth, timeout=60)
            response.raise_for_status()
            content = response.content

        # Step 2: Parse Excel and extract tracking numbers
        workbook = load_workbook(filename=io.BytesIO(content), data_only=False)
        sheet = workbook.active

        tracking_data = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1), start=2):
            cell_a = row[0]
            if not cell_a.value:
                continue

            cell_value = str(cell_a.value)
            digits_only = re.sub(r'\D', '', cell_value)

            if len(digits_only) != 12:
                warning_msg = f"Row {idx}: Invalid tracking number '{cell_value}' - Expected 12 digits. Skipping."
                warnings.append(warning_msg)
                logger.warning(f"[Task {task_id}] {warning_msg}")
                continue

            tracking_data.append((idx, digits_only))

        logger.info(f"[Task {task_id}] Extracted {len(tracking_data)} valid tracking numbers")

        # Step 3: 构造 URLs（每 10 个追踪号一个 URL）
        url_data = []

        for i in range(0, len(tracking_data), 10):
            batch_chunk = tracking_data[i:i+10]

            # 在当前批次内去重
            seen = set()
            unique_chunk = []
            for row_num, tracking_num in batch_chunk:
                if tracking_num not in seen:
                    seen.add(tracking_num)
                    unique_chunk.append((row_num, tracking_num))
                else:
                    warning_msg = f"Row {row_num}: Duplicate tracking number in batch. Skipping."
                    warnings.append(warning_msg)
                    logger.warning(f"[Task {task_id}] {warning_msg}")

            if not unique_chunk:
                continue

            start_row = unique_chunk[0][0]
            end_row = unique_chunk[-1][0]

            # 构造 URL
            params = {}
            for j in range(1, 11):
                if j <= len(unique_chunk):
                    params[f'requestNo{j}'] = unique_chunk[j-1][1]
                else:
                    params[f'requestNo{j}'] = ''

            params['search.x'] = str(random.randint(1, 173))
            params['search.y'] = str(random.randint(1, 45))
            params['startingUrlPatten'] = ''
            params['locale'] = 'ja'

            base_url = 'https://trackings.post.japanpost.jp/services/srv/search'
            url = f"{base_url}?{urlencode(params)}"
            custom_id_suffix = f"{start_row}-{end_row}"

            url_data.append((url, custom_id_suffix, start_row))

        logger.info(f"[Task {task_id}] Constructed {len(url_data)} URLs")

        # Step 4: Create or find TrackingBatch
        batch, created = TrackingBatch.objects.get_or_create(
            file_path=file_path,
            task_name=task_name,
            defaults={
                'batch_uuid': uuid.uuid4(),
                'celery_task_id': task_id,
                'total_jobs': len(url_data),
                'status': 'pending'
            }
        )

        batch_uuid_str = str(batch.batch_uuid)
        batch_short = batch_uuid_str[:8]

        if created:
            logger.info(f"[Task {task_id}] Created TrackingBatch {batch_short}")
        else:
            logger.info(f"[Task {task_id}] Found existing TrackingBatch {batch_short}")

        # Step 5: 批量投递任务到 publish_tracking_queue
        dispatched_count = 0
        skipped_count = 0

        for idx, (url, custom_id_suffix, start_row) in enumerate(url_data):
            custom_id = f"{config['custom_id_prefix']}-{batch_short}-{custom_id_suffix}"

            # 检查是否已投递
            if TrackingJob.objects.filter(batch=batch, custom_id=custom_id).exists():
                skipped_count += 1
                logger.debug(f"[Task {task_id}] Skipping already dispatched: {custom_id}")
                continue

            # 投递任务
            publish_tracking_batch.apply_async(
                args=[task_name, url, batch_uuid_str, custom_id, start_row],
                countdown=dispatched_count * 2  # 每个任务延迟 2 秒
            )

            dispatched_count += 1

        logger.info(
            f"[Task {task_id}] Dispatched {dispatched_count} tasks, skipped {skipped_count}"
        )

        # 记录 SyncLog
        SyncLog.objects.create(
            operation_type=config['sync_log_triggered'],
            celery_task_id=task_id,
            message=f"Dispatched {dispatched_count} publish tasks",
            success=True,
            details={
                'file_path': file_path,
                'batch_uuid': batch_uuid_str,
                'total_urls': len(url_data),
                'dispatched': dispatched_count,
                'skipped': skipped_count,
                'warnings': warnings
            }
        )

        return {
            'status': 'success',
            'batch_uuid': batch_uuid_str,
            'total_urls': len(url_data),
            'dispatched': dispatched_count,
            'skipped': skipped_count,
            'warnings': warnings
        }

    except Exception as exc:
        logger.error(
            f"[Task {task_id}] {config['display_name']} task failed: {exc}",
            exc_info=True
        )

        # 记录失败日志
        SyncLog.objects.create(
            operation_type=config['sync_log_triggered'],
            celery_task_id=task_id,
            message=f"Failed to process {file_path}: {str(exc)}",
            success=False,
            details={
                'file_path': file_path,
                'error': str(exc)
            }
        )

        # Retry if retries remaining
        if self.request.retries < self.max_retries:
            raise self.retry(exc=exc)
        raise
  
  
  


# ============================================================================
# Yamato Tracking 10 - Direct Query Implementation
# ============================================================================

def query_yamato(tracking_numbers):
    """
    查询大和运输追踪信息（批量查询，最多10个）
    
    Args:
        tracking_numbers: 追踪号列表，最多10个
        
    Returns:
        requests.Response 对象
    """
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.ssl_ import create_urllib3_context
    import ssl
    
    # 自定义SSL上下文
    class CustomHTTPAdapter(HTTPAdapter):
        def init_poolmanager(self, *args, **kwargs):
            ctx = create_urllib3_context(
                ssl_version=ssl.PROTOCOL_TLS,  # 自动选择最佳版本
                ciphers='DEFAULT'
            )
            # 禁用某些检查
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_REQUIRED
            kwargs['ssl_context'] = ctx
            return super().init_poolmanager(*args, **kwargs)
    
    # 使用自定义适配器
    session = requests.Session()
    session.mount('https://', CustomHTTPAdapter())
    
    url = "https://toi.kuronekoyamato.co.jp/cgi-bin/tneko"
    data = {
        "mypagesession": "",
        "backaddress": "",
        "backrequest": "get",
    }
    
    # 填充追踪号（最多10个）
    for i, number in enumerate(tracking_numbers[:10], 1):
        data[f"number{i:02d}"] = number
    for i in range(len(tracking_numbers) + 1, 11):
        data[f"number{i:02d}"] = ""
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'ja-JP,ja;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Referer': 'https://toi.kuronekoyamato.co.jp/cgi-bin/tneko',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    }
    
    try:
        response = session.post(
            url,
            data=data,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return response
    except Exception as e:
        logger.error(f"Query yamato error: {e}")
        raise


@app.task(
    name='apps.data_acquisition.tasks.process_yamato_tracking_10_excel',
    bind=True,
    max_retries=2,
    default_retry_delay=300,
    queue='tracking_excel_queue'
)
def process_yamato_tracking_10_excel(self, file_path, document_url=None):
    """
    处理 Yamato Tracking 10 任务
    
    此任务不使用 WebScraper API，而是直接调用 query_yamato() 函数进行批量查询。
    
    流程：
    1. 下载 Excel 文件
    2. 提取 A 列追踪号（从第 2 行开始）
    3. 创建 TrackingBatch 和 TrackingJob（支持断点续传）
    4. 每 10 个追踪号调用一次 query_yamato()
    5. 保存查询结果状态码到 TrackingJob.writeback_data
    6. 整组标记为 completed 或 failed
    
    Args:
        file_path: Nextcloud 文件路径
        document_url: OnlyOffice 提供的下载 URL（可选）
        
    Returns:
        dict: 处理结果统计
    """
    from .models import SyncLog, TrackingBatch, TrackingJob
    from django.conf import settings
    from django.utils import timezone
    import uuid
    import random
    
    task_name = 'yamato_tracking_10'
    config = TRACKING_TASK_CONFIGS[task_name]
    task_id = self.request.id
    
    logger.info(
        f"[Task {task_id}] Starting {config['display_name']} for {file_path}"
    )
    
    try:
        # ============================================================================
        # Step 1: 下载 Excel 文件
        # ============================================================================
        content = None
        
        if document_url:
            logger.info(f"[Task {task_id}] Downloading from OnlyOffice URL...")
            response = requests.get(document_url, timeout=30)
            response.raise_for_status()
            content = response.content
        else:
            logger.info(f"[Task {task_id}] Downloading from WebDAV...")
            from apps.data_aggregation.nextcloud_client import get_nextcloud_client
            nc_client = get_nextcloud_client()
            content = nc_client.download_file(file_path)
        
        if not content:
            raise ValueError("Failed to download Excel file")
        
        # ============================================================================
        # Step 2: 解析 Excel，提取 A 列追踪号
        # ============================================================================
        logger.info(f"[Task {task_id}] Parsing Excel file...")
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        tracking_numbers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1), start=2):
            cell_a = row[0]
            if cell_a.value:
                tracking_number = str(cell_a.value).strip()
                if tracking_number:
                    tracking_numbers.append({
                        'number': tracking_number,
                        'row_index': row_idx - 2  # 从 0 开始计数
                    })
        
        logger.info(
            f"[Task {task_id}] Extracted {len(tracking_numbers)} tracking numbers"
        )
        
        if not tracking_numbers:
            logger.warning(f"[Task {task_id}] No tracking numbers found in Excel")
            return {
                "status": "success",
                "message": "No tracking numbers found",
                "total": 0,
                "processed": 0
            }
        
        # ============================================================================
        # Step 3: 查找或创建 TrackingBatch（支持断点续传）
        # ============================================================================
        tracking_batch = TrackingBatch.objects.filter(
            file_path=file_path,
            task_name=task_name,
            status__in=['pending', 'processing']
        ).order_by('-created_at').first()
        
        if tracking_batch:
            # 断点续传：查询已创建的 TrackingJob
            existing_indices = set(TrackingJob.objects.filter(
                batch=tracking_batch
            ).values_list('index', flat=True))
            
            batch_uuid_str = str(tracking_batch.batch_uuid)
            batch_short = batch_uuid_str[:8]
            
            logger.info(
                f"[Task {task_id}] Resuming batch {batch_short}, "
                f"existing jobs: {len(existing_indices)}"
            )
        else:
            # 创建新批次
            batch_uuid = uuid.uuid4()
            batch_uuid_str = str(batch_uuid)
            batch_short = batch_uuid_str[:8]

            # 计算 TrackingJob 数量（每 10 个追踪号一个 job）
            total_jobs = (len(tracking_numbers) + 9) // 10

            tracking_batch = TrackingBatch.objects.create(
                batch_uuid=batch_uuid,
                task_name=task_name,
                file_path=file_path,
                celery_task_id=task_id,
                total_jobs=total_jobs,
                status='pending'
            )

            existing_indices = set()

            logger.info(
                f"[Task {task_id}] Created new batch {batch_short} "
                f"for {len(tracking_numbers)} tracking numbers ({total_jobs} jobs)"
            )
        
        # 记录触发 SyncLog
        SyncLog.objects.create(
            operation_type=config['sync_log_triggered'],
            celery_task_id=task_id,
            file_path=file_path,
            message=f"{config['display_name']} task triggered",
            success=True,
            details={
                'batch_uuid': batch_uuid_str,
                'total_numbers': len(tracking_numbers)
            }
        )
        
        # ============================================================================
        # Step 4: 创建 TrackingJob（每 10 个追踪号创建一个 TrackingJob）
        # ============================================================================
        query_url = "https://toi.kuronekoyamato.co.jp/cgi-bin/tneko"
        jobs_to_create = []

        # 将追踪号按每 10 个一组分批
        for batch_index in range(0, len(tracking_numbers), 10):
            batch_items = tracking_numbers[batch_index:batch_index+10]

            # 检查这一批是否已经有 TrackingJob 创建（用于断点续传）
            # 使用批次索引作为 index
            job_index = batch_index // 10
            if job_index in existing_indices:
                continue

            # 构建 target_url: 查询URL（单号1｜单号2｜...）
            batch_numbers = [item['number'] for item in batch_items]
            tracking_numbers_str = "｜".join(batch_numbers)
            target_url = f"{query_url}（{tracking_numbers_str}）"

            custom_id = f"{config['custom_id_prefix']}-{batch_short}-{job_index:04d}"
            jobs_to_create.append(TrackingJob(
                batch=tracking_batch,
                custom_id=custom_id,
                target_url=target_url,
                index=job_index,
                status='pending'
            ))

        if jobs_to_create:
            TrackingJob.objects.bulk_create(jobs_to_create)
            logger.info(
                f"[Task {task_id}] Created {len(jobs_to_create)} TrackingJobs (each for up to 10 tracking numbers)"
            )
        
        # ============================================================================
        # Step 5: 批量查询（每个 TrackingJob 对应一批 10 个追踪号）
        # ============================================================================
        # 注意：必须在循环前将 QuerySet 物化为 list，否则在循环中修改 status 后，
        # 使用 QuerySet 切片会导致 OFFSET 计算错误，跳过部分 job
        all_jobs_list = list(TrackingJob.objects.filter(
            batch=tracking_batch,
            status='pending'
        ).order_by('index'))

        total_jobs = len(all_jobs_list)
        processed_count = 0
        success_count = 0
        failed_count = 0

        logger.info(
            f"[Task {task_id}] Starting batch query for {total_jobs} pending jobs"
        )

        # 遍历每个 TrackingJob（每个 job 已经包含一批追踪号）
        for i, tracking_job in enumerate(all_jobs_list):
            # 从 target_url 中解析出追踪号
            # 格式：https://toi.kuronekoyamato.co.jp/cgi-bin/tneko（单号1｜单号2｜...）
            target_url = tracking_job.target_url
            if '（' in target_url and '）' in target_url:
                numbers_part = target_url.split('（')[1].split('）')[0]
                batch_numbers = numbers_part.split('｜')
            else:
                # 兼容旧格式（如果有的话）
                logger.warning(f"[Task {task_id}] Unexpected target_url format: {target_url}")
                batch_numbers = [target_url]


            batch_num = i + 1
            total_batches = total_jobs
            progress_pct = (processed_count / total_jobs * 100) if total_jobs > 0 else 0

            logger.info(
                f"[Task {task_id}] Processing job {batch_num}/{total_batches} "
                f"({len(batch_numbers)} numbers) - Progress: {progress_pct:.1f}%"
            )

            try:
                # 调用 query_yamato
                response = query_yamato(batch_numbers)
                status_code = response.status_code

                logger.info(
                    f"[Task {task_id}] Batch {batch_num}/{total_batches} query successful - "
                    f"Status: {status_code}, Numbers: {batch_numbers}"
                )

                # ============================================================================
                # 解析HTML并落库
                # ============================================================================
                from .yamato_parser import extract_tracking_data
                from apps.data_aggregation.models import Purchasing
                from datetime import datetime as dt
                import re

                html_content = response.text
                tracking_data = extract_tracking_data(html_content, year=2026)

                logger.info(
                    f"[Task {task_id}] Batch {batch_num}/{total_batches} extracted {len(tracking_data)} tracking records from HTML"
                )

                batch_updated_count = 0
                batch_skipped_count = 0
                batch_error_count = 0

                for data in tracking_data:
                    # 跳过没有tracking_number的记录
                    if not data.get('tracking_number'):
                        logger.warning(
                            f"[Task {task_id}] Batch {batch_num}/{total_batches} skipping record with no tracking_number: {data}"
                        )
                        batch_skipped_count += 1
                        continue

                    tracking_num = data['tracking_number']
                    delivery_date = data.get('delivery_date')
                    delivery_status = data.get('delivery_status')

                    # 使用正则提取数字部分进行匹配
                    digits_only = re.sub(r'\D', '', tracking_num)

                    # 查询匹配的Purchasing实例
                    matching_records = Purchasing.objects.filter(
                        tracking_number__regex=r'.*'.join(digits_only)
                    )

                    count = matching_records.count()

                    if count == 0:
                        logger.error(
                            f"[Task {task_id}] Batch {batch_num}/{total_batches} no Purchasing instance found for tracking_number: {tracking_num}"
                        )
                        batch_error_count += 1
                        continue
                    elif count > 1:
                        logger.error(
                            f"[Task {task_id}] Batch {batch_num}/{total_batches} multiple ({count}) Purchasing instances found for tracking_number: {tracking_num}. "
                            f"Skipping update to avoid ambiguity."
                        )
                        batch_error_count += 1
                        continue

                    # 找到唯一实例，执行更新
                    purchasing_instance = matching_records.first()
                    current_time = timezone.now()

                    # 准备更新字段
                    update_kwargs = {
                        'last_info_updated_at': current_time,
                        'updated_at': current_time,
                        'delivery_status_query_source': 'process_yamato_tracking_10_excel',
                    }

                    # 将delivery_date转换为datetime
                    if delivery_date:
                        delivery_datetime = dt.combine(delivery_date, dt.min.time())
                        delivery_datetime = timezone.make_aware(delivery_datetime)
                        update_kwargs['delivery_status_query_time'] = delivery_datetime

                    # 更新delivery_status
                    if delivery_status:
                        update_kwargs['latest_delivery_status'] = delivery_status

                    try:
                        purchasing_instance.update_fields(**update_kwargs)
                        batch_updated_count += 1
                        logger.info(
                            f"[Task {task_id}] Batch {batch_num}/{total_batches} successfully updated Purchasing {purchasing_instance.uuid} "
                            f"(tracking: {tracking_num}, status: {delivery_status}, date: {delivery_date})"
                        )
                    except Exception as e:
                        logger.error(
                            f"[Task {task_id}] Batch {batch_num}/{total_batches} failed to update Purchasing {purchasing_instance.uuid}: {e}",
                            exc_info=True
                        )
                        batch_error_count += 1

                logger.info(
                    f"[Task {task_id}] Batch {batch_num}/{total_batches} database update complete - "
                    f"Updated: {batch_updated_count}, Skipped: {batch_skipped_count}, Errors: {batch_error_count}"
                )

                # 标记 TrackingJob 为成功
                tracking_job.writeback_data = str(status_code)
                tracking_job.status = 'completed'
                tracking_job.completed_at = timezone.now()
                tracking_job.save()

                success_count += 1

            except Exception as e:
                logger.error(
                    f"[Task {task_id}] Batch {batch_num}/{total_batches} query failed: {e}",
                    exc_info=True
                )

                # 标记 TrackingJob 为失败
                tracking_job.status = 'failed'
                tracking_job.error_message = str(e)
                tracking_job.completed_at = timezone.now()
                tracking_job.save()

                failed_count += 1

            processed_count += 1
            
            # 更新批次进度
            tracking_batch.update_progress()
            
            # 随机睡眠 1-50 秒
            if processed_count < total_jobs:  # 最后一批不需要睡眠
                sleep_time = random.randint(1, 50)
                progress_after = (processed_count / total_jobs * 100) if total_jobs > 0 else 0
                logger.info(
                    f"[Task {task_id}] Progress: {progress_after:.1f}% "
                    f"({processed_count}/{total_jobs} jobs) - Sleeping {sleep_time}s before next batch"
                )
                time.sleep(sleep_time)
        
        # ============================================================================
        # Step 6: 完成处理
        # ============================================================================
        tracking_batch.update_progress()
        
        logger.info(
            f"[Task {task_id}] Batch {batch_short} complete: "
            f"total={total_jobs}, success={success_count}, failed={failed_count}"
        )
        
        # 记录完成 SyncLog
        SyncLog.objects.create(
            operation_type=config['sync_log_completed'],
            celery_task_id=task_id,
            file_path=file_path,
            message=f"Processed {processed_count} tracking numbers",
            success=True,
            details={
                'batch_uuid': batch_uuid_str,
                'total': total_jobs,
                'success': success_count,
                'failed': failed_count,
                'tracking_batch_id': tracking_batch.id
            }
        )
        
        return {
            "status": "success",
            "batch_uuid": batch_uuid_str,
            "file_path": file_path,
            "total": total_jobs,
            "success": success_count,
            "failed": failed_count,
            "tracking_batch_id": tracking_batch.id
        }
        
    except Exception as exc:
        logger.error(
            f"[Task {task_id}] {config['display_name']} task failed: {exc}",
            exc_info=True
        )


        
        # 记录失败 SyncLog
        SyncLog.objects.create(
            operation_type=config['sync_log_completed'],
            celery_task_id=task_id,
            file_path=file_path,
            message=f"Task failed: {str(exc)}",
            success=False,
            error_message=str(exc)
        )
        
        # Retry if retries remaining
        if self.request.retries < self.max_retries:
            raise self.retry(exc=exc)
        raise



# ============================================================================
# Yamato Tracking 10 Tracking Number - Query from Purchasing Model
# ============================================================================

@app.task(
    name='apps.data_acquisition.tasks.process_yamato_tracking_10_tracking_number',
    bind=True,
    max_retries=0,  # 不重试
    queue='yamato_tracking_10_tracking_number_queue'
)
def process_yamato_tracking_10_tracking_number(self):
    """
    从Purchasing模型查询符合条件的记录并进行Yamato物流查询

    查询条件：
    1. order_number以w开头（不区分大小写）
    2. shipping_method是"YAMATO TRANSPORT CO.,LTD."
    3. tracking_number中提取的数字是12位
    4. latest_delivery_status不是"配達完了"或"お届け先にお届け済み"
    5. 如果latest_delivery_status是"＊＊ お問い合わせ番号が見つかりません..."，忽略时间限制

    最多查询10条记录，使用query_yamato()进行批量查询。

    Returns:
        dict: 处理结果统计
    """
    from apps.data_aggregation.models import Purchasing
    from .models import SyncLog
    from django.utils import timezone
    from django.db.models import Q
    import re
    import random
    
    task_id = self.request.id
    
    logger.info(
        f"[Task {task_id}] Starting Yamato Tracking 10 Tracking Number query from Purchasing model"
    )
    
    try:
        # ============================================================================
        # Step 1: 从Purchasing模型查询符合条件的记录
        # ============================================================================
        
        # 获取查询间隔配置（默认6小时）
        from django.conf import settings
        from datetime import timedelta

        query_interval_hours = getattr(settings, 'DELIVERY_STATUS_QUERY_INTERVAL_HOURS', 6)
        time_threshold = timezone.now() - timedelta(hours=query_interval_hours)

        # 查询条件：
        # 1. order_number以w开头（不区分大小写）
        # 2. shipping_method是"YAMATO TRANSPORT CO.,LTD."
        # 3. latest_delivery_status不是"配達完了"或"お届け先にお届け済み"
        # 4. 防止重复发布：last_info_updated_at 为空或超过配置的时间间隔
        purchasing_records = Purchasing.objects.filter(
            Q(order_number__istartswith='w'),
            Q(shipping_method='YAMATO TRANSPORT CO.,LTD.'),
            ~Q(latest_delivery_status__in=['配達完了', 'お届け先にお届け済み']),
            Q(last_info_updated_at__isnull=True) | Q(last_info_updated_at__lt=time_threshold)
        ).exclude(
            tracking_number__isnull=True
        ).exclude(
            tracking_number=''
        )[:100]  # 先取100条，后续筛选
        
        logger.info(
            f"[Task {task_id}] Found {len(purchasing_records)} candidate records"
        )
        
        # ============================================================================
        # Step 2: 使用正则表达式筛选tracking_number
        # ============================================================================
        valid_records = []
        for record in purchasing_records:
            # 提取tracking_number中的所有数字
            digits = re.sub(r'\D', '', record.tracking_number)

            # 判断是否为12位
            if len(digits) == 12:
                valid_records.append({
                    'record': record,
                    'tracking_number': digits,
                    'uuid': record.uuid,
                    'order_number': record.order_number
                })

                # 最多取10条
                if len(valid_records) >= 10:
                    break

        logger.info(
            f"[Task {task_id}] Filtered {len(valid_records)} valid records with 12-digit tracking numbers"
        )
        
        if not valid_records:
            logger.warning(f"[Task {task_id}] No valid records found")
            
            SyncLog.objects.create(
                operation_type='yamato_tracking_10_tracking_number_completed',
                celery_task_id=task_id,
                message="No valid records found",
                success=True,
                details={
                    'total_candidates': len(purchasing_records),
                    'valid_records': 0
                }
            )
            
            return {
                "status": "success",
                "message": "No valid records found",
                "total_candidates": len(purchasing_records),
                "valid_records": 0,
                "processed": 0
            }
        
        # ============================================================================
        # Step 3: 创建 TrackingBatch 和 TrackingJob
        # ============================================================================
        from .models import TrackingBatch, TrackingJob
        import uuid as uuid_module

        batch_uuid = uuid_module.uuid4()
        batch_uuid_str = str(batch_uuid)
        batch_short = batch_uuid_str[:8]

        tracking_batch = TrackingBatch.objects.create(
            batch_uuid=batch_uuid,
            task_name='yamato_tracking_10_tracking_number',
            file_path='',  # 该任务不基于文件
            celery_task_id=task_id,
            total_jobs=1,  # 只创建一个 job
            status='pending'
        )

        logger.info(
            f"[Task {task_id}] Created TrackingBatch {batch_short} for {len(valid_records)} tracking numbers"
        )

        # 构建 target_url: 查询URL（单号1｜单号2｜...）
        tracking_numbers = [item['tracking_number'] for item in valid_records]
        query_url = "https://toi.kuronekoyamato.co.jp/cgi-bin/tneko"
        tracking_numbers_str = "｜".join(tracking_numbers)
        target_url = f"{query_url}（{tracking_numbers_str}）"

        custom_id = f"yttn-{batch_short}-0001"
        tracking_job = TrackingJob.objects.create(
            batch=tracking_batch,
            custom_id=custom_id,
            target_url=target_url,
            index=0,
            status='pending'
        )

        logger.info(
            f"[Task {task_id}] Created TrackingJob {custom_id} with target_url: {target_url}"
        )

        # 记录触发 SyncLog
        SyncLog.objects.create(
            operation_type='yamato_tracking_10_tracking_number_triggered',
            celery_task_id=task_id,
            message=f"Yamato Tracking 10 Tracking Number task triggered",
            success=True,
            details={
                'batch_uuid': batch_uuid_str,
                'total_numbers': len(tracking_numbers),
                'tracking_numbers': tracking_numbers
            }
        )

        # ============================================================================
        # Step 4: 调用query_yamato()进行批量查询
        # ============================================================================
        
        logger.info(
            f"[Task {task_id}] Querying Yamato with {len(tracking_numbers)} tracking numbers: {tracking_numbers}"
        )
        
        # 随机睡眠 1-50 毫秒（反节流）
        sleep_time = random.randint(1, 50) / 1000.0
        logger.info(f"[Task {task_id}] Sleeping {sleep_time*1000:.1f}ms before query")
        import time
        time.sleep(sleep_time)
        
        try:
            response = query_yamato(tracking_numbers)
            status_code = response.status_code

            logger.info(
                f"[Task {task_id}] Query successful - Status: {status_code}"
            )

            # 更新 TrackingJob 状态为 completed
            tracking_job.writeback_data = str(status_code)
            tracking_job.status = 'completed'
            tracking_job.completed_at = timezone.now()
            tracking_job.save()

            # 更新 TrackingBatch 进度
            tracking_batch.update_progress()

            logger.info(
                f"[Task {task_id}] Updated TrackingJob {custom_id} status to completed"
            )

            # ============================================================================
            # Step 5: 解析HTML并落库
            # ============================================================================
            from .yamato_parser import extract_tracking_data
            from datetime import datetime as dt

            html_content = response.text
            tracking_data = extract_tracking_data(html_content, year=2026)

            logger.info(
                f"[Task {task_id}] Extracted {len(tracking_data)} tracking records from HTML"
            )

            updated_count = 0
            skipped_count = 0
            error_count = 0

            for data in tracking_data:
                # 跳过没有tracking_number的记录
                if not data.get('tracking_number'):
                    logger.warning(
                        f"[Task {task_id}] Skipping record with no tracking_number: {data}"
                    )
                    skipped_count += 1
                    continue

                tracking_num = data['tracking_number']
                delivery_date = data.get('delivery_date')
                delivery_status = data.get('delivery_status')

                # 使用正则提取数字部分进行匹配
                digits_only = re.sub(r'\D', '', tracking_num)

                # 查询匹配的Purchasing实例
                matching_records = Purchasing.objects.filter(
                    tracking_number__regex=r'.*'.join(digits_only)
                )

                count = matching_records.count()

                if count == 0:
                    logger.error(
                        f"[Task {task_id}] No Purchasing instance found for tracking_number: {tracking_num}"
                    )
                    error_count += 1
                    continue
                elif count > 1:
                    logger.error(
                        f"[Task {task_id}] Multiple ({count}) Purchasing instances found for tracking_number: {tracking_num}. "
                        f"Skipping update to avoid ambiguity."
                    )
                    error_count += 1
                    continue

                # 找到唯一实例，执行更新
                purchasing_instance = matching_records.first()
                current_time = timezone.now()

                # 准备更新字段
                update_kwargs = {
                    'last_info_updated_at': current_time,
                    'updated_at': current_time,
                    'delivery_status_query_source': 'process_yamato_tracking_10_tracking_number',
                }

                # 将delivery_date转换为datetime
                if delivery_date:
                    delivery_datetime = dt.combine(delivery_date, dt.min.time())
                    delivery_datetime = timezone.make_aware(delivery_datetime)
                    update_kwargs['delivery_status_query_time'] = delivery_datetime

                # 更新delivery_status
                if delivery_status:
                    update_kwargs['latest_delivery_status'] = delivery_status

                try:
                    purchasing_instance.update_fields(**update_kwargs)

                    # 解锁记录
                    purchasing_instance.is_locked = False
                    purchasing_instance.locked_at = None
                    purchasing_instance.locked_by_worker = 'unlock'
                    purchasing_instance.save(update_fields=['is_locked', 'locked_at', 'locked_by_worker'])

                    updated_count += 1
                    logger.info(
                        f"[Task {task_id}] Successfully updated Purchasing {purchasing_instance.uuid} "
                        f"(tracking: {tracking_num}, status: {delivery_status}, date: {delivery_date}, unlocked=True)"
                    )
                except Exception as e:
                    logger.error(
                        f"[Task {task_id}] Failed to update Purchasing {purchasing_instance.uuid}: {e}",
                        exc_info=True
                    )
                    error_count += 1

            logger.info(
                f"[Task {task_id}] Database update complete - "
                f"Updated: {updated_count}, Skipped: {skipped_count}, Errors: {error_count}"
            )
            
            # 记录成功日志
            SyncLog.objects.create(
                operation_type='yamato_tracking_10_tracking_number_completed',
                celery_task_id=task_id,
                message=f"Successfully queried {len(tracking_numbers)} tracking numbers and updated {updated_count} records",
                success=True,
                details={
                    'batch_uuid': batch_uuid_str,
                    'total_candidates': len(purchasing_records),
                    'valid_records': len(valid_records),
                    'processed': len(tracking_numbers),
                    'status_code': status_code,
                    'tracking_numbers': tracking_numbers,
                    'order_numbers': [item['order_number'] for item in valid_records],
                    'extracted_records': len(tracking_data),
                    'updated_count': updated_count,
                    'skipped_count': skipped_count,
                    'error_count': error_count,
                    'tracking_batch_id': tracking_batch.id
                }
            )

            return {
                "status": "success",
                "batch_uuid": batch_uuid_str,
                "total_candidates": len(purchasing_records),
                "valid_records": len(valid_records),
                "processed": len(tracking_numbers),
                "status_code": status_code,
                "tracking_numbers": tracking_numbers,
                "extracted_records": len(tracking_data),
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "error_count": error_count,
                "tracking_batch_id": tracking_batch.id
            }
            
        except Exception as e:
            logger.error(
                f"[Task {task_id}] Query failed: {e}",
                exc_info=True
            )

            # 更新 TrackingJob 状态为 failed
            tracking_job.status = 'failed'
            tracking_job.error_message = str(e)
            tracking_job.completed_at = timezone.now()
            tracking_job.save()

            # 更新 TrackingBatch 进度
            tracking_batch.update_progress()

            logger.info(
                f"[Task {task_id}] Updated TrackingJob {custom_id} status to failed"
            )

            # 记录失败日志（不重试）
            SyncLog.objects.create(
                operation_type='yamato_tracking_10_tracking_number_completed',
                celery_task_id=task_id,
                message=f"Query failed: {str(e)}",
                success=False,
                error_message=str(e),
                details={
                    'batch_uuid': batch_uuid_str,
                    'total_candidates': len(purchasing_records),
                    'valid_records': len(valid_records),
                    'tracking_numbers': tracking_numbers,
                    'order_numbers': [item['order_number'] for item in valid_records],
                    'tracking_batch_id': tracking_batch.id
                }
            )

            return {
                "status": "failed",
                "batch_uuid": batch_uuid_str,
                "total_candidates": len(purchasing_records),
                "valid_records": len(valid_records),
                "error": str(e),
                "tracking_numbers": tracking_numbers,
                "tracking_batch_id": tracking_batch.id
            }
        
    except Exception as exc:
        logger.error(
            f"[Task {task_id}] Task execution failed: {exc}",
            exc_info=True
        )
        
        # 记录失败日志
        SyncLog.objects.create(
            operation_type='yamato_tracking_10_tracking_number_completed',
            celery_task_id=task_id,
            message=f"Task execution failed: {str(exc)}",
            success=False,
            error_message=str(exc)
        )
        
        return {
            "status": "error",
            "error": str(exc)
        }

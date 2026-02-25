"""
Excel parser for reading Nextcloud Excel files.
Parses DATA sheet with __id, __version, __op columns.
"""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Custom exception for Excel parsing errors."""
    pass


class ExcelParser:
    """
    Parser for Nextcloud Excel files with DATA sheet.

    Expected format:
    Sheet "DATA":
    | __id | __version | __op | field1 | field2 | ... |
    |  123 | 2025-01-15 10:30:00 | UPDATE | value1 | value2 | ... |
    |      | 2025-01-15 10:35:00 |        | value1 | value2 | ... |  <- New record
    |  456 | 2025-01-15 10:40:00 | DELETE |        |        | ... |
    """

    REQUIRED_COLUMNS = ['__id', '__version', '__op']
    SHEET_NAME = 'DATA'

    def __init__(self, excel_bytes: bytes):
        """
        Initialize parser with Excel file content.

        Args:
            excel_bytes: Excel file content as bytes

        Raises:
            ExcelParseError: If file cannot be loaded
        """
        try:
            self.workbook = load_workbook(BytesIO(excel_bytes), data_only=True)
        except InvalidFileException as e:
            raise ExcelParseError(f"Invalid Excel file: {e}")
        except Exception as e:
            raise ExcelParseError(f"Error loading Excel file: {e}")

        # Validate DATA sheet exists
        if self.SHEET_NAME not in self.workbook.sheetnames:
            raise ExcelParseError(
                f"Missing '{self.SHEET_NAME}' sheet. "
                f"Available sheets: {', '.join(self.workbook.sheetnames)}"
            )

        self.sheet = self.workbook[self.SHEET_NAME]
        self.headers = self._parse_headers()
        self._validate_headers()

    def _parse_headers(self) -> List[str]:
        """
        Parse header row (first row).

        Returns:
            List of column names
        """
        headers = []
        first_row = list(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        for cell_value in first_row:
            if cell_value is None:
                break  # Stop at first empty column
            headers.append(str(cell_value).strip())

        return headers

    def _validate_headers(self):
        """
        Validate that required columns exist.

        Raises:
            ExcelParseError: If required columns are missing
        """
        missing_columns = set(self.REQUIRED_COLUMNS) - set(self.headers)
        if missing_columns:
            raise ExcelParseError(
                f"Missing required columns: {', '.join(missing_columns)}. "
                f"Found columns: {', '.join(self.headers)}"
            )

    def _get_column_index(self, column_name: str) -> int:
        """Get 0-based index of column by name."""
        try:
            return self.headers.index(column_name)
        except ValueError:
            raise ExcelParseError(f"Column '{column_name}' not found")

    def _parse_cell_value(self, value: Any, field_name: str) -> Any:
        """
        Parse cell value to appropriate Python type.

        Args:
            value: Raw cell value
            field_name: Field name for logging

        Returns:
            Parsed value
        """
        if value is None or value == '':
            return None

        # Handle datetime objects
        if isinstance(value, datetime):
            return value

        # Handle date objects
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        # Handle strings
        if isinstance(value, str):
            value = value.strip()
            if value == '':
                return None

        return value

    def parse_rows(self) -> List[Dict[str, Any]]:
        """
        Parse all data rows from DATA sheet.

        Returns:
            List of dictionaries, each representing a row with column names as keys

        Example:
            [
                {
                    '__id': 123,
                    '__version': '2025-01-15 10:30:00',
                    '__op': 'UPDATE',
                    'order_number': 'ORD-001',
                    'confirmed_at': datetime(2025, 1, 15, 10, 0, 0),
                    ...
                },
                {
                    '__id': None,  # New record
                    '__version': '2025-01-15 10:35:00',
                    '__op': None,
                    'order_number': 'ORD-002',
                    ...
                }
            ]
        """
        rows = []
        row_num = 1  # Track row number for error messages

        # Iterate from row 2 (skip header)
        for excel_row in self.sheet.iter_rows(min_row=2, values_only=True):
            row_num += 1

            # Check if row is empty (first cell is None)
            if excel_row[0] is None and excel_row[1] is None:
                continue  # Skip empty rows

            # Parse row into dictionary
            row_data = {}
            for col_idx, header in enumerate(self.headers):
                if col_idx >= len(excel_row):
                    row_data[header] = None
                else:
                    raw_value = excel_row[col_idx]
                    row_data[header] = self._parse_cell_value(raw_value, header)

            # Parse __id (can be empty for new records)
            id_value = row_data.get('__id')
            if id_value is not None:
                try:
                    row_data['__id'] = int(id_value)
                except (ValueError, TypeError):
                    logger.warning(
                        f"Row {row_num}: Invalid __id value '{id_value}', treating as new record"
                    )
                    row_data['__id'] = None

            # Parse __version (required)
            version_value = row_data.get('__version')
            if not version_value:
                logger.warning(f"Row {row_num}: Missing __version, skipping row")
                continue

            # Convert __version to string for comparison
            if isinstance(version_value, datetime):
                row_data['__version'] = version_value.isoformat()
            else:
                row_data['__version'] = str(version_value)

            # Parse __op (optional: UPDATE, DELETE, or empty for CREATE)
            op_value = row_data.get('__op')
            if op_value:
                row_data['__op'] = str(op_value).strip().upper()
            else:
                row_data['__op'] = None

            # Store original row number for error reporting
            row_data['_row_num'] = row_num

            rows.append(row_data)

        logger.info(f"Parsed {len(rows)} rows from Excel file")
        return rows

    def get_field_names(self) -> List[str]:
        """
        Get list of business field names (excluding __id, __version, __op).

        Returns:
            List of field names
        """
        return [
            header for header in self.headers
            if header not in self.REQUIRED_COLUMNS
        ]

    def close(self):
        """Close workbook."""
        if self.workbook:
            self.workbook.close()

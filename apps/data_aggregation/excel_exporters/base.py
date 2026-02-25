"""
Base Excel exporter class for data_aggregation models.
数据聚合模型的基础Excel导出器类。
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.apps import apps


class BaseExcelExporter:
    """
    Base class for Excel exporters.
    All model-specific exporters should inherit from this class.
    """
    # Model name to export (must be set by subclasses)
    model_name = None

    # App label (default is data_aggregation)
    app_label = 'data_aggregation'

    def __init__(self):
        if not self.model_name:
            raise NotImplementedError("model_name must be set in subclass")

    def get_model(self):
        """
        Get the Django model class.

        Returns:
            Model: Django model class

        Raises:
            LookupError: If model is not found
        """
        try:
            return apps.get_model(self.app_label, self.model_name)
        except LookupError:
            raise ValueError(f"Model '{self.model_name}' not found in {self.app_label} app")

    def get_queryset(self):
        """
        Get the queryset for the model.
        Override this method to customize the queryset (e.g., filtering, ordering).

        Returns:
            QuerySet: Django queryset
        """
        model = self.get_model()
        if hasattr(model, 'is_deleted'):
            return model.objects.filter(is_deleted=False)
        return model.objects.all()

    def get_fields(self):
        """
        Get the fields to export.
        Override this method to customize which fields are exported.

        Returns:
            list: List of field names to export
        """
        model = self.get_model()
        fields = []
        for field in model._meta.get_fields():
            if field.many_to_many or field.one_to_many:
                continue
            fields.append(field.name)
            if hasattr(field, 'remote_field') and field.remote_field:
                fields.append(f"{field.name}_id")
        return fields

    def get_header_names(self):
        """
        Get the header names for the Excel file.
        Override this method to customize header names.

        Returns:
            dict: Dictionary mapping field names to header names
        """
        # By default, use field names as headers
        return {field: field for field in self.get_fields()}

    def format_cell_value(self, obj, field_name):
        """
        Format a cell value for export.
        Override this method to customize how values are formatted.

        Args:
            obj: Model instance
            field_name (str): Name of the field

        Returns:
            Formatted value for the cell
        """
        value = getattr(obj, field_name, None)

        # Handle different field types
        if value is None:
            return ''
        elif hasattr(value, 'pk'):  # Foreign key
            return str(value)
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, dict):
            return str(value)
        else:
            return value

    def customize_workbook(self, wb, ws):
        """
        Customize the workbook after data is written.
        Override this method to add custom styling, formulas, etc.

        Args:
            wb: openpyxl Workbook object
            ws: openpyxl Worksheet object
        """
        # Default implementation: auto-adjust column widths
        for col_num in range(1, len(self.get_fields()) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

    def export(self):
        """
        Export the model data to Excel format.

        Returns:
            io.BytesIO: Excel file as bytes stream
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.model_name

        # Get fields and headers
        fields = self.get_fields()
        header_names = self.get_header_names()

        # Write header row
        for col_num, field_name in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_names.get(field_name, field_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Get data
        queryset = self.get_queryset()

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(fields, 1):
                cell_value = self.format_cell_value(obj, field_name)
                ws.cell(row=row_num, column=col_num, value=cell_value)

        # Allow subclasses to customize the workbook
        self.customize_workbook(wb, ws)

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

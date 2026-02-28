from playwright.sync_api import sync_playwright

def automated_parcel_tracking(url):
    with sync_playwright() as p:
        # 启动 Chromium 浏览器（可替换为 firefox 或 webkit）
        browser = p.chromium.launch(headless=False)  # headless=False 显示浏览器窗口
        page = browser.new_page()
        page.goto(url)
        print(page.title())
        browser.close()


def excel_hyperlink():

    from openpyxl import load_workbook
    import re

    path = "some-data/OWRYT-20260108-1108.xlsx"
    wb = load_workbook(path, data_only=True)  # data_only=True: 取计算结果(如有公式)
    ws = wb["工作表1"]  # 或 wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:  # 存在单元格超链接
                text = cell.value
                url = cell.hyperlink.target  # 超链接 URL
                # 有些文件会把地址放在 location（内部跳转），也可以一并看
                loc = cell.hyperlink.location

                m = re.search(
                    r'(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])', url)
                email = m.group(1) if m else None
                print(cell.coordinate, text, url, email)


def read_csv():
    import pandas as pd
    path = "some-data/official_website_redirect_to_yamato_tracking-4.csv"
    df=pd.read_csv(path)

    print(df)





if __name__ == '__main__':
    read_csv()